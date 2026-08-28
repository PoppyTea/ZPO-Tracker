"""
Porządkowanie powtórzeń w arkuszu przed importem.

Narzędzie na czas przejściowy: miesiące zaległości są sklejone z pracy
kilku osób i ten sam odbiór bywa wpisany dwa razy. Do czasu, aż zrobi to
auto-naprawa w potoku importu, robi to jedno polecenie na pliku.

Reguła wzięta z pomiaru na realnym sierpniu, nie z wyobraźni. Odległości
między powtórzeniami układały się tam tak:

    1,1,1,1,1,1,1,1,2,2,2,3,4,4,4,4,7,8,9   |   11,12,31,47,51,53,80,245,247,3429,3431

Przerwa między 9 a 11 jest naturalną granicą i stąd próg dziesięciu
wierszy. Poniżej to jeden blok arkusza — ktoś rozbił jeden odbiór na dwa
wiersze, więc **ilości się sumują**. Powyżej to dwie osoby wpisujące
w dwóch miejscach — tego **nie wolno rozstrzygać automatycznie**, bo nie
wiadomo, czy to dwa odbiory, czy dwa opisy jednego.

Decyzje podjęte przez Papavera: blisko → sumuj, daleko → oznacz i zostaw.
Nazw firm narzędzie NIE RUSZA.
"""
import openpyxl
import pytest

from zpo_tracker import porzadkowanie


NAGLOWKI = ["data", "Kurier", " Pełna Nazwa Nadawcy",
            "Adres odbioru dla wszystkich nadawców",
            " Wpisujemy łączną liczbę odebranych Pocztexów", "PNI ZPO"]


@pytest.fixture
def arkusz(tmp_path):
    def zbuduj(wiersze):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sierpień"
        ws.append(NAGLOWKI)
        for w in wiersze:
            ws.append(w)
        sciezka = tmp_path / "zrodlo.xlsx"
        wb.save(sciezka)
        return sciezka
    return zbuduj


def _w(ilosc=1, nadawca="Żabka", adres="Kwiatowa 8", kurier="Kowalski Jan",
       data="2026-08-03", pni=None):
    return [data, kurier, nadawca, adres, ilosc, pni]


def _wczytaj(sciezka):
    ws = openpyxl.load_workbook(sciezka).active
    return [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]


# --- blisko: scalanie z sumowaniem --------------------------------------

def test_sasiadujace_powtorzenie_scala_sie_i_sumuje_ilosci(arkusz, tmp_path):
    wynik = tmp_path / "out.xlsx"
    raport = porzadkowanie.porzadkuj(arkusz([_w(ilosc=2), _w(ilosc=4)]), wynik)

    wiersze = _wczytaj(wynik)
    assert len(wiersze) == 1
    assert wiersze[0][4] == 6
    assert raport.scalone == 1


def test_scalanie_ratuje_pole_wypelnione_tylko_w_jednym_wierszu(arkusz, tmp_path):
    """PNI wpisane w drugim wierszu, a w pierwszym puste — po scaleniu
    musi zostać. Inaczej porządkowanie gubiłoby dane, zamiast je łączyć."""
    wynik = tmp_path / "out.xlsx"
    porzadkowanie.porzadkuj(arkusz([_w(ilosc=2), _w(ilosc=4, pni="228648")]), wynik)
    assert _wczytaj(wynik)[0][5] == "228648"


def test_prog_dziesieciu_wierszy_jest_granica(arkusz, tmp_path):
    """Dokładnie na progu jeszcze scalamy, wiersz dalej już nie."""
    wypelniacz = [_w(adres=f"Inna {i}") for i in range(9)]
    wynik = tmp_path / "out.xlsx"
    r = porzadkowanie.porzadkuj(arkusz([_w(ilosc=2)] + wypelniacz + [_w(ilosc=4)]), wynik)
    assert (r.scalone, r.do_decyzji) == (1, 0)

    wynik2 = tmp_path / "out2.xlsx"
    r2 = porzadkowanie.porzadkuj(
        arkusz([_w(ilosc=2)] + wypelniacz + [_w(adres="X 1")] + [_w(ilosc=4)]), wynik2)
    assert (r2.scalone, r2.do_decyzji) == (0, 1)


# --- daleko: oznaczenie bez ruszania ------------------------------------

def test_odlegle_powtorzenie_zostaje_i_jest_oznaczone(arkusz, tmp_path):
    """Dwie osoby wpisujące w dwóch miejscach - nie wiadomo, czy to dwa
    odbiory, czy dwa opisy jednego. Automat nie ma prawa zgadywać."""
    wypelniacz = [_w(adres=f"Inna {i}") for i in range(30)]
    wynik = tmp_path / "out.xlsx"
    r = porzadkowanie.porzadkuj(arkusz([_w(ilosc=2)] + wypelniacz + [_w(ilosc=4)]), wynik)

    wiersze = _wczytaj(wynik)
    assert len(wiersze) == 32           # nic nie usunięto
    assert r.do_decyzji == 1
    oznaczone = [w for w in wiersze if w[-1] and "DO DECYZJI" in str(w[-1])]
    assert len(oznaczone) == 2


def test_uwaga_wskazuje_numer_drugiego_wiersza(arkusz, tmp_path):
    """Bez numeru użytkownik szuka pary wzrokiem po pięciu tysiącach
    wierszy."""
    wypelniacz = [_w(adres=f"Inna {i}") for i in range(30)]
    wynik = tmp_path / "out.xlsx"
    porzadkowanie.porzadkuj(arkusz([_w()] + wypelniacz + [_w()]), wynik)
    uwagi = [str(w[-1]) for w in _wczytaj(wynik) if w[-1]]
    assert any("33" in u for u in uwagi)


# --- niezmienniki, które muszą trzymać ----------------------------------

def test_sumy_ilosci_sa_ZACHOWANE(arkusz, tmp_path):
    """
    NAJWAŻNIEJSZY test tego pliku. Narzędzie zmienia liczbę wierszy, więc
    jedyną rzeczą, która musi zostać nietknięta, są sumy - to one idą do
    rozliczenia. Sam złapałem się na błędzie w tej kontroli, licząc sumę
    „przed" na już zmodyfikowanych danych; stąd sprawdzenie z osobnego
    odczytu pliku źródłowego.
    """
    zrodlo = arkusz([_w(ilosc=2), _w(ilosc=4), _w(ilosc=7, adres="Inna 1")])
    wynik = tmp_path / "out.xlsx"
    raport = porzadkowanie.porzadkuj(zrodlo, wynik)

    przed = sum(w[4] for w in _wczytaj(zrodlo) if isinstance(w[4], int))
    po = sum(w[4] for w in _wczytaj(wynik) if isinstance(w[4], int))
    assert przed == po == 13
    assert raport.sumy_zgodne is True


def test_narzedzie_NIE_RUSZA_nazw_firm(arkusz, tmp_path):
    """Jawne polecenie Papavera. Dwie pisownie to dla tego narzędzia dwa
    różne punkty - scalanie nazw to osobna decyzja i osobne narzędzie."""
    wynik = tmp_path / "out.xlsx"
    porzadkowanie.porzadkuj(
        arkusz([_w(nadawca="Żabka"), _w(nadawca="żabka")]), wynik)

    nazwy = sorted(w[2] for w in _wczytaj(wynik))
    assert nazwy == ["Żabka", "żabka"]


def test_plik_bez_powtorzen_przechodzi_bez_zmian(arkusz, tmp_path):
    wynik = tmp_path / "out.xlsx"
    r = porzadkowanie.porzadkuj(
        arkusz([_w(adres="A 1"), _w(adres="B 2")]), wynik)
    assert (r.scalone, r.do_decyzji, len(_wczytaj(wynik))) == (0, 0, 2)


def test_rozne_dni_to_nie_powtorzenie(arkusz, tmp_path):
    """Klucz zawiera datę - ten sam punkt obsłużony w dwa dni to dwa
    odrębne odbiory, nie duplikat."""
    wynik = tmp_path / "out.xlsx"
    r = porzadkowanie.porzadkuj(
        arkusz([_w(data="2026-08-03"), _w(data="2026-08-04")]), wynik)
    assert (r.scalone, r.do_decyzji) == (0, 0)


def test_rozni_kurierzy_to_nie_powtorzenie(arkusz, tmp_path):
    wynik = tmp_path / "out.xlsx"
    r = porzadkowanie.porzadkuj(
        arkusz([_w(kurier="Kowalski Jan"), _w(kurier="Nowak Anna")]), wynik)
    assert (r.scalone, r.do_decyzji) == (0, 0)
