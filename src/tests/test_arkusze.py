"""
Jedna warstwa czytania skoroszytów, niezależna od formatu.

Powód istnienia: `openpyxl` **nie otwiera `.xls` w ogóle** - to inny
format pliku (OLE2), nie kwestia wersji. Wszystkie eksporty z BaŚKi
przychodzą właśnie w nim, więc bez tej warstwy nie da się wczytać ani
rejonarza, ani rejestru punktów ZPO.

Dwie rzeczy, na których stoi cały ten moduł i które mają tu testy:

1. **Format rozpoznajemy po ZAWARTOŚCI, nie po rozszerzeniu.** Pliki
   przychodzą przez przeglądarkę i bywają przemianowane; rozszerzenie
   jest deklaracją użytkownika, magiczne bajty są faktem.

2. **Oba silniki muszą oddawać TE SAME wartości.** `xlrd` zwraca każdą
   liczbę jako `float`, więc numer domu `8` przychodzi jako `8.0` i po
   sklejeniu klucza daje `"8.0"` zamiast `"8"` - czyli inny adres.
   Ujednolicenie jest tutaj, żeby żaden konsument nie musiał o tym
   pamiętać.
"""
import pytest
import xlwt
from openpyxl import Workbook

from zpo_tracker import arkusze


@pytest.fixture
def plik_xlsx(tmp_path):
    def zbuduj(nazwa="dane.xlsx", arkusze_dane=None):
        arkusze_dane = arkusze_dane or {"Arkusz1": [["a", "b"], [1, 2]]}
        wb = Workbook()
        wb.remove(wb.active)
        for tytul, wiersze in arkusze_dane.items():
            ws = wb.create_sheet(tytul)
            for w in wiersze:
                ws.append(w)
        sciezka = tmp_path / nazwa
        wb.save(sciezka)
        return sciezka
    return zbuduj


@pytest.fixture
def plik_xls(tmp_path):
    def zbuduj(nazwa="dane.xls", arkusze_dane=None):
        arkusze_dane = arkusze_dane or {"Arkusz1": [["a", "b"], [1, 2]]}
        wb = xlwt.Workbook()
        for tytul, wiersze in arkusze_dane.items():
            ws = wb.add_sheet(tytul)
            for i, wiersz in enumerate(wiersze):
                for j, wartosc in enumerate(wiersz):
                    ws.write(i, j, wartosc)
        sciezka = tmp_path / nazwa
        wb.save(str(sciezka))
        return sciezka
    return zbuduj


# --- oba formaty czytają się tak samo -----------------------------------

def test_xlsx_daje_nazwy_arkuszy_i_wiersze(plik_xlsx):
    with arkusze.otworz(plik_xlsx()) as sk:
        assert sk.nazwy_arkuszy() == ["Arkusz1"]
        assert list(sk.wiersze("Arkusz1")) == [("a", "b"), (1, 2)]


def test_xls_daje_nazwy_arkuszy_i_wiersze(plik_xls):
    with arkusze.otworz(plik_xls()) as sk:
        assert sk.nazwy_arkuszy() == ["Arkusz1"]
        assert list(sk.wiersze("Arkusz1")) == [("a", "b"), (1, 2)]


def test_wiele_arkuszy_w_kolejnosci_z_pliku(plik_xls):
    sciezka = plik_xls(arkusze_dane={
        "WA87": [["Miejscowość"], ["Warszawa"]],
        "L11": [["Miejscowość"], ["Legionowo"]],
        "RDH": [["Miejscowość"], ["Radzymin"]],
    })
    with arkusze.otworz(sciezka) as sk:
        assert sk.nazwy_arkuszy() == ["WA87", "L11", "RDH"]


# --- rozszerzenie kłamie -----------------------------------------------

def test_xls_pod_nazwa_xlsx_otwiera_sie_poprawnie(plik_xls):
    """Realny scenariusz: plik przychodzi przez przeglądarkę i ktoś
    dopisuje mu rozszerzenie „żeby Excel go otworzył". Zaufanie nazwie
    dawałoby tu błąd formatu zamiast danych."""
    sciezka = plik_xls(nazwa="podstepny.xlsx")
    with arkusze.otworz(sciezka) as sk:
        assert list(sk.wiersze("Arkusz1")) == [("a", "b"), (1, 2)]


def test_xlsx_pod_nazwa_xls_otwiera_sie_poprawnie(plik_xlsx):
    sciezka = plik_xlsx(nazwa="podstepny.xls")
    with arkusze.otworz(sciezka) as sk:
        assert list(sk.wiersze("Arkusz1")) == [("a", "b"), (1, 2)]


def test_plik_ktory_nie_jest_arkuszem_daje_czytelny_blad(tmp_path):
    sciezka = tmp_path / "notatka.xlsx"
    sciezka.write_text("to nie jest arkusz", encoding="utf-8")
    with pytest.raises(arkusze.NieznanyFormat) as e:
        arkusze.otworz(sciezka)
    assert "notatka.xlsx" in str(e.value)


# --- ujednolicenie wartości --------------------------------------------

def test_liczba_calkowita_jest_intem_w_obu_silnikach(plik_xls, plik_xlsx):
    """PUŁAPKA, dla której ta warstwa w ogóle ujednolica typy: `xlrd`
    oddaje każdą liczbę jako float, więc numer domu `8` przychodzi jako
    `8.0`. Sklejony w klucz daje `"8.0"` - czyli inny adres niż `"8"`,
    i rejonarz przestaje odpowiadać dla całej klasy wierszy."""
    dane = {"A": [["nr"], [8]]}
    with arkusze.otworz(plik_xls(arkusze_dane=dane)) as sk:
        z_xls = list(sk.wiersze("A"))[1]
    with arkusze.otworz(plik_xlsx(arkusze_dane=dane)) as sk:
        z_xlsx = list(sk.wiersze("A"))[1]
    assert z_xls == z_xlsx == (8,)


def test_liczba_ulamkowa_zostaje_ulamkiem(plik_xls):
    with arkusze.otworz(plik_xls(arkusze_dane={"A": [["x"], [1.5]]})) as sk:
        assert list(sk.wiersze("A"))[1] == (1.5,)


def test_pusty_arkusz_daje_pusty_iterator(plik_xls):
    with arkusze.otworz(plik_xls(arkusze_dane={"Puste": []})) as sk:
        assert list(sk.wiersze("Puste")) == []


# --- pamięć: właściwe ryzyko przy 219 arkuszach -------------------------

def test_arkusz_jest_zwalniany_po_przeczytaniu(plik_xls):
    """Realny rejonarz to 219 arkuszy i 21 MB. `xlrd` bez zwalniania
    trzyma wszystkie wczytane naraz - i to jest jedyny powód, dla którego
    ten import mógłby się nie zmieścić w pamięci na słabszej stacji.
    Test pilnuje kontraktu, nie zużycia: po przejściu arkusza ma on być
    zwolniony."""
    sciezka = plik_xls(arkusze_dane={
        "A": [["x"], [1]], "B": [["x"], [2]]})
    with arkusze.otworz(sciezka) as sk:
        list(sk.wiersze("A"))
        assert sk.zwolnione() == ["A"]
        list(sk.wiersze("B"))
        assert sorted(sk.zwolnione()) == ["A", "B"]


def test_zwalnianie_nie_psuje_ponownego_odczytu(plik_xls):
    """Zwolnienie ma być niewidoczne dla wołającego - drugi przebieg po
    tym samym arkuszu musi dać to samo, a nie pustkę."""
    with arkusze.otworz(plik_xls()) as sk:
        pierwszy = list(sk.wiersze("Arkusz1"))
        drugi = list(sk.wiersze("Arkusz1"))
    assert pierwszy == drugi == [("a", "b"), (1, 2)]
