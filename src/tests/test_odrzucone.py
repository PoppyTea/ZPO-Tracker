"""
Raport odrzuconych wierszy importu: `.xlsx` ze wszystkim, co nie weszło,
plus powód i NUMER WIERSZA w pliku źródłowym.

Numer wiersza jest tu sednem, nie ozdobą. "71 wierszy wymagało uwagi"
bez wskazania, których, jest informacją, z którą nie da się nic zrobić;
z numerem to jest lista zadań do poprawienia w źródłowym Excelu.

Powód istnienia (Papaver, 2026-08-24): import ma dawać opcję "pomiń
niespójności" z osobnym zatwierdzeniem, a ta opcja bez wykazu tego, co
pominięto, byłaby po prostu cichą utratą danych z ładniejszą nazwą.
"""
from datetime import date

import openpyxl

from zpo_tracker import eksport, import_orchestrator
from zpo_tracker.models import WierszImportu


def _surowy(numer=None, **nadpisz):
    dane = {
        "data": date(2026, 8, 3),
        " Pełna Nazwa Nadawcy": "Żabka",
        "Adres odbioru dla wszystkich nadawców": "Odkryta 24",
        "Kurier": "Kowalski Jan",
        "Rejon": "WA87",
        " Wpisujemy łączną liczbę odebranych Pocztexów": 3,
        "PNI ZPO": "228648",
    }
    dane.update(nadpisz)
    if numer is not None:
        dane[import_orchestrator.KLUCZ_NUMERU_WIERSZA] = numer
    return dane


def _wczytaj(sciezka):
    ws = openpyxl.load_workbook(sciezka).active
    naglowki = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    wiersze = [dict(zip(naglowki, w)) for w in ws.iter_rows(min_row=2, values_only=True)]
    return naglowki, wiersze


# --- zbieranie z dwóch źródeł ------------------------------------------

def test_zbiera_odrzucone_z_walidacji():
    pozycje = import_orchestrator.zbierz_odrzucone(
        [{"wiersz": _surowy(numer=7), "powod": "zła data"}])
    assert len(pozycje) == 1
    assert pozycje[0]["numer_wiersza"] == 7
    assert pozycje[0]["powod"] == "zła data"
    assert pozycje[0]["dane"]["Kurier"] == "Kowalski Jan"


def test_zbiera_takze_wiersze_wymagajace_uwagi():
    """Duplikaty i konflikty PNI/adres wychodzą dopiero przy zapisie,
    więc niosą już zwalidowany obiekt, nie surowy dict."""
    w = WierszImportu(data=date(2026, 8, 3), nadawca="Żabka", adres="Odkryta 24",
                      kurier="Kowalski Jan", rejon="WA87", ilosc_total=3)
    pozycje = import_orchestrator.zbierz_odrzucone(
        [], [{"wiersz": w, "powod": "duplikat"}])
    assert pozycje[0]["powod"] == "duplikat"
    assert pozycje[0]["dane"]["kurier"] == "Kowalski Jan"


def test_brak_numeru_wiersza_nie_wybucha():
    """Konflikty wykryte przy zapisie nie znają numeru wiersza - lepiej
    raport bez numeru niż brak raportu."""
    pozycje = import_orchestrator.zbierz_odrzucone(
        [{"wiersz": _surowy(), "powod": "cokolwiek"}])
    assert pozycje[0]["numer_wiersza"] is None


def test_klucz_numeru_nie_wycieka_do_danych():
    """`__wiersz__` to metadana transportowa, nie kolumna źródłowa -
    w raporcie ma być osobną kolumną, nie doklejonym śmieciem."""
    pozycje = import_orchestrator.zbierz_odrzucone(
        [{"wiersz": _surowy(numer=3), "powod": "x"}])
    assert import_orchestrator.KLUCZ_NUMERU_WIERSZA not in pozycje[0]["dane"]


# --- zapis do .xlsx -----------------------------------------------------

def test_zapisuje_plik_z_numerem_i_powodem(tmp_path):
    sciezka = tmp_path / "odrzucone.xlsx"
    ile = eksport.zapisz_odrzucone(sciezka, [
        {"numer_wiersza": 12, "powod": "brak daty", "dane": _surowy()},
    ])
    assert ile == 1
    naglowki, wiersze = _wczytaj(sciezka)
    assert naglowki[:2] == [eksport.NAGLOWEK_NUMERU, eksport.NAGLOWEK_POWODU]
    assert wiersze[0][eksport.NAGLOWEK_NUMERU] == 12
    assert wiersze[0][eksport.NAGLOWEK_POWODU] == "brak daty"


def test_zachowuje_oryginalne_naglowki_i_wartosci(tmp_path):
    """Plik ma być poprawialny obok źródła, więc kolumny muszą wyglądać
    tak samo - łącznie ze spacją w ' Pełna Nazwa Nadawcy', która jest
    częścią danych, nie literówką."""
    sciezka = tmp_path / "o.xlsx"
    eksport.zapisz_odrzucone(sciezka, [
        {"numer_wiersza": 5, "powod": "x", "dane": _surowy()},
    ])
    naglowki, wiersze = _wczytaj(sciezka)
    assert " Pełna Nazwa Nadawcy" in naglowki
    assert wiersze[0][" Pełna Nazwa Nadawcy"] == "Żabka"
    assert wiersze[0]["PNI ZPO"] == "228648"


def test_pni_zostaje_tekstem(tmp_path):
    """Ta sama pułapka co w eksporcie miesiąca: '007' rzutowane na int
    to '7', czyli inny punkt."""
    sciezka = tmp_path / "o.xlsx"
    eksport.zapisz_odrzucone(sciezka, [
        {"numer_wiersza": 1, "powod": "x", "dane": _surowy(**{"PNI ZPO": "007"})},
    ])
    _, wiersze = _wczytaj(sciezka)
    assert wiersze[0]["PNI ZPO"] == "007"


def test_rozne_zestawy_kolumn_daja_sume_kolumn(tmp_path):
    """Odrzucenia przychodzą z dwóch źródeł o różnych kształtach -
    raport musi pomieścić oba, nie zgubić kolumn drugiego."""
    sciezka = tmp_path / "o.xlsx"
    eksport.zapisz_odrzucone(sciezka, [
        {"numer_wiersza": 1, "powod": "a", "dane": {"Kurier": "Jan"}},
        {"numer_wiersza": 2, "powod": "b", "dane": {"Rejon": "WA87"}},
    ])
    naglowki, wiersze = _wczytaj(sciezka)
    assert "Kurier" in naglowki and "Rejon" in naglowki
    assert wiersze[0]["Rejon"] is None
    assert wiersze[1]["Kurier"] is None


def test_pusta_lista_pisze_sam_naglowek(tmp_path):
    """Plik powstaje nawet bez odrzuceń - "nic nie odrzucono" ma być
    widocznym artefaktem, nie brakiem pliku, którego nie wiadomo jak
    zinterpretować."""
    sciezka = tmp_path / "o.xlsx"
    assert eksport.zapisz_odrzucone(sciezka, []) == 0
    naglowki, wiersze = _wczytaj(sciezka)
    assert naglowki == [eksport.NAGLOWEK_NUMERU, eksport.NAGLOWEK_POWODU]
    assert wiersze == []


def test_kolejnosc_wierszy_jak_w_zrodle(tmp_path):
    sciezka = tmp_path / "o.xlsx"
    eksport.zapisz_odrzucone(sciezka, [
        {"numer_wiersza": n, "powod": "x", "dane": {"Kurier": f"K{n}"}}
        for n in (9, 3, 14)
    ])
    _, wiersze = _wczytaj(sciezka)
    assert [w[eksport.NAGLOWEK_NUMERU] for w in wiersze] == [9, 3, 14]


# --- przepływ end-to-end ------------------------------------------------

def test_walidacja_niesie_numer_wiersza_do_raportu(tmp_path):
    """Cały łańcuch: surowe wiersze z numerami -> walidacja -> zbierz ->
    plik. To jest ta ścieżka, która ma dać pewność przy imporcie."""
    surowe = [
        _surowy(numer=2),
        _surowy(numer=3, **{" Wpisujemy łączną liczbę odebranych Pocztexów": "aaa"}),
    ]
    _, odrzucone = import_orchestrator.zwaliduj_wiersze(surowe)
    assert len(odrzucone) == 1

    sciezka = tmp_path / "o.xlsx"
    eksport.zapisz_odrzucone(sciezka, import_orchestrator.zbierz_odrzucone(odrzucone))
    _, wiersze = _wczytaj(sciezka)
    assert wiersze[0][eksport.NAGLOWEK_NUMERU] == 3


def test_numer_wiersza_nie_psuje_walidacji():
    """Metadana nie ma prawa wejść do WierszImportu ani go wywrócić."""
    zwalidowane, odrzucone = import_orchestrator.zwaliduj_wiersze([_surowy(numer=2)])
    assert len(zwalidowane) == 1 and not odrzucone
    assert zwalidowane[0].kurier == "Kowalski Jan"


# --- odczyt pliku w GUI -------------------------------------------------

def test_odczyt_pliku_dokleja_numery_wierszy(tmp_path):
    """Klasa błędu, której nie złapie żadne czytanie kodu: helper GUI
    odwoływał się do modułu, który był zaimportowany tylko po nazwach,
    więc wywaliłby się dopiero przy realnym imporcie pliku - czyli
    u użytkownika, nie u nas."""
    from zpo_tracker.gui.zakladka_import_export import _wczytaj_surowe_wiersze

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Kurier", "Rejon"])
    ws.append(["Kowalski Jan", "WA87"])
    ws.append(["Nowak Anna", "WA88"])
    sciezka = tmp_path / "zrodlo.xlsx"
    wb.save(sciezka)

    wiersze = _wczytaj_surowe_wiersze(sciezka)
    # Numeracja liczona tak, jak widzi ją człowiek w Excelu: nagłówek to
    # wiersz 1, więc pierwsze dane są w wierszu 2.
    assert [w[import_orchestrator.KLUCZ_NUMERU_WIERSZA] for w in wiersze] == [2, 3]
    assert wiersze[0]["Kurier"] == "Kowalski Jan"


# --- plik-reszta: kopia oryginału bez zaimportowanych wierszy ----------
#
# Papaver 2026-08-24: wygodniejsze od raportu, bo nie trzeba wyłuskiwać
# wierszy - dostajesz ten sam plik, tylko krótszy o to, co weszło.
# Poprawiasz i importujesz ponownie ten sam plik.

NAGLOWKI_ZRODLA = ["data", "Kurier", " Pełna Nazwa Nadawcy",
                   "Adres odbioru dla wszystkich nadawców", "Rejon",
                   " Wpisujemy łączną liczbę odebranych Pocztexów", "PNI ZPO"]


def test_reszta_ma_dokladnie_naglowki_zrodla(tmp_path):
    sciezka = tmp_path / "reszta.xlsx"
    eksport.zapisz_niezaimportowane(
        sciezka, NAGLOWKI_ZRODLA, [_surowy(numer=3)], z_powodem=False)
    naglowki, _ = _wczytaj(sciezka)
    assert naglowki == NAGLOWKI_ZRODLA


def test_reszta_nie_przecieka_metadana_numeru(tmp_path):
    sciezka = tmp_path / "reszta.xlsx"
    eksport.zapisz_niezaimportowane(
        sciezka, NAGLOWKI_ZRODLA, [_surowy(numer=3)], z_powodem=False)
    naglowki, _ = _wczytaj(sciezka)
    assert import_orchestrator.KLUCZ_NUMERU_WIERSZA not in naglowki


def test_reszta_zachowuje_wartosci(tmp_path):
    sciezka = tmp_path / "reszta.xlsx"
    eksport.zapisz_niezaimportowane(
        sciezka, NAGLOWKI_ZRODLA, [_surowy(numer=3)], z_powodem=False)
    _, wiersze = _wczytaj(sciezka)
    assert wiersze[0][" Pełna Nazwa Nadawcy"] == "Żabka"
    assert wiersze[0]["PNI ZPO"] == "228648"


def test_reszta_moze_dopisac_powod_na_koncu(tmp_path):
    """Domyślnie dopisujemy powód jako OSTATNIĄ kolumnę - przy ponownym
    imporcie i tak zostanie zignorowana (mapowanie filtruje po znanych
    nagłówkach), a bez niej użytkownik nie wie, co poprawić."""
    sciezka = tmp_path / "reszta.xlsx"
    eksport.zapisz_niezaimportowane(
        sciezka, NAGLOWKI_ZRODLA, [_surowy(numer=3)], powody={3: "duplikat"})
    naglowki, wiersze = _wczytaj(sciezka)
    assert naglowki[-1] == eksport.NAGLOWEK_POWODU
    assert wiersze[0][eksport.NAGLOWEK_POWODU] == "duplikat"


def test_wybiera_tylko_niezaimportowane():
    surowe = [_surowy(numer=2), _surowy(numer=3), _surowy(numer=4)]
    pozycje = [{"numer_wiersza": 3, "powod": "x", "dane": {}}]
    reszta = import_orchestrator.wybierz_niezaimportowane(surowe, pozycje)
    assert [w[import_orchestrator.KLUCZ_NUMERU_WIERSZA] for w in reszta] == [3]


def test_konflikt_przy_zapisie_tez_trafia_do_reszty():
    """Duplikaty wychodzą dopiero przy zapisie i niosą WierszImportu -
    numer wiersza musi przez niego przejść, inaczej te wiersze zniknęłyby
    z pliku-reszty, czyli po cichu przepadły."""
    surowe = [_surowy(numer=2), _surowy(numer=5)]
    zwalidowane, _ = import_orchestrator.zwaliduj_wiersze(surowe)
    pozycje = import_orchestrator.zbierz_odrzucone(
        [], [{"wiersz": zwalidowane[1], "powod": "duplikat"}])
    assert pozycje[0]["numer_wiersza"] == 5
    reszta = import_orchestrator.wybierz_niezaimportowane(surowe, pozycje)
    assert len(reszta) == 1


def test_numer_wiersza_nie_wchodzi_do_danych_raportu():
    surowe = [_surowy(numer=5)]
    zwalidowane, _ = import_orchestrator.zwaliduj_wiersze(surowe)
    pozycje = import_orchestrator.zbierz_odrzucone(
        [], [{"wiersz": zwalidowane[0], "powod": "duplikat"}])
    assert "numer_wiersza" not in pozycje[0]["dane"]


def test_plik_reszta_daje_sie_zaimportowac_ponownie(tmp_path):
    """Zamknięcie pętli: poprawiasz plik-resztę i wrzucasz go z powrotem.
    Gdyby kolumny albo typy się rozjechały, ta ścieżka by tego nie
    przeżyła - a to jest dokładnie ten scenariusz, dla którego plik
    powstaje."""
    from zpo_tracker.gui.zakladka_import_export import _wczytaj_surowe_wiersze

    sciezka = tmp_path / "reszta.xlsx"
    eksport.zapisz_niezaimportowane(
        sciezka, NAGLOWKI_ZRODLA, [_surowy(numer=3)], powody={3: "duplikat"})

    ponownie = _wczytaj_surowe_wiersze(sciezka)
    zwalidowane, odrzucone = import_orchestrator.zwaliduj_wiersze(ponownie)
    assert not odrzucone
    assert zwalidowane[0].kurier == "Kowalski Jan"
    assert zwalidowane[0].pni_zpo == "228648"
