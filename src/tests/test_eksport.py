"""
Testy exportu do .xlsx: układ kolumn identyczny ze snapshotem źródłowym,
plus test wierności round-tripu na realnej próbce danych. TDD.
"""
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from zpo_tracker import repo, eksport
from zpo_tracker.eksport import nazwa_arkusza
from zpo_tracker.importer import get_or_create_punkt, import_row
from zpo_tracker.normalizacja import REJON_NIEZNANY

REALNA_PROBKA = (
    Path(__file__).resolve().parent.parent.parent
    / "data" / "real-data-samples" / "2026-08-07-snapshot-ZPO.xlsx"
)


@pytest.fixture
def conn():
    conn = repo.polacz(":memory:")
    repo.utworz_schemat(conn)
    yield conn
    conn.close()


def _wstaw_prosta_transakcje(conn, **nadpisz):
    dane = dict(
        data="2026-08-03", kurier_id=None, punkt_id=None, rejon_id=None,
    )
    dane.update(nadpisz)
    kurier_id = conn.execute(
        "INSERT INTO kurierzy (imie_nazwisko) VALUES (?)", ("Kowalski Jan",)
    ).lastrowid
    punkt_id, _ = get_or_create_punkt(conn, "Żabka", "Odkryta 24", "228648")
    conn.execute(
        """INSERT INTO transakcje (data, kurier_id, punkt_id, ilosc_total, ilosc_zpo)
           VALUES (?, ?, ?, ?, ?)""",
        (dane["data"], kurier_id, punkt_id, 3, 3),
    )


def test_naglowki_identyczne_ze_snapshotem():
    assert eksport.NAGLOWKI[0] == "data"
    assert eksport.NAGLOWKI[3] == "Kurier"
    assert eksport.NAGLOWKI[7] == "PNI ZPO"
    assert eksport.NAGLOWKI[-1] == "Wykonawca"
    assert len(eksport.NAGLOWKI) == 13


def test_nazwa_arkusza_po_polsku():
    assert eksport.nazwa_arkusza(2026, 8) == "Sierpień"
    assert eksport.nazwa_arkusza(2026, 6) == "Czerwiec"


def test_eksportuj_miesiac_zapisuje_plik_z_poprawnym_arkuszem(conn, tmp_path):
    _wstaw_prosta_transakcje(conn)
    sciezka = tmp_path / "export.xlsx"
    liczba = eksport.eksportuj_miesiac(conn, 2026, 8, sciezka)
    assert liczba == 1

    wb = openpyxl.load_workbook(sciezka)
    assert wb.sheetnames == ["Sierpień"]
    ws = wb["Sierpień"]
    naglowki = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert naglowki == eksport.NAGLOWKI

    wiersz = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
    assert wiersz[0].date() == date(2026, 8, 3)  # openpyxl zawsze odczytuje jako datetime
    assert wiersz[3] == "Kowalski Jan"
    assert wiersz[5] == 3  # ilosc_total jako int
    # PNI jako TEKST (0.1-alpha.3.2) - to klucz tożsamości punktu, a nie
    # liczba: rzutowanie na int gubiło zera wiodące i rozdwajało punkt
    assert wiersz[7] == "228648"


def test_eksportuj_miesiac_pomija_inne_miesiace(conn, tmp_path):
    _wstaw_prosta_transakcje(conn, data="2026-07-15")
    sciezka = tmp_path / "export.xlsx"
    liczba = eksport.eksportuj_miesiac(conn, 2026, 8, sciezka)
    assert liczba == 0


@pytest.mark.skipif(
    not REALNA_PROBKA.exists(),
    reason="wymaga realnej próbki danych (gitignored, patrz data/README.md)",
)
def test_round_trip_import_export_na_realnych_danych(conn, tmp_path):
    """
    Import realnego snapshotu -> export -> odczyt exportu -> porównanie
    liczby wierszy i typów komórek. Jedyny wiarygodny dowód wierności
    round-tripu (patrz plan MVP).
    """
    wb_zrodlo = openpyxl.load_workbook(REALNA_PROBKA, data_only=True)
    ws_zrodlo = wb_zrodlo["Czerwiec"]
    naglowki = [c.value for c in next(ws_zrodlo.iter_rows(min_row=1, max_row=1))]

    zaimportowano = 0
    for wartosci in ws_zrodlo.iter_rows(min_row=2, values_only=True):
        wiersz = dict(zip(naglowki, wartosci))
        wynik = import_row(conn, wiersz)
        if not wynik["skipped"]:
            zaimportowano += 1

    assert zaimportowano > 1000  # rząd wielkości z realnej próbki (1239 z 4573)

    sciezka = tmp_path / "export.xlsx"
    liczba_wyeksportowanych = eksport.eksportuj_miesiac(conn, 2026, 8, sciezka)
    assert liczba_wyeksportowanych == zaimportowano

    wb_export = openpyxl.load_workbook(sciezka)
    ws_export = wb_export["Sierpień"]
    naglowki_export = [c.value for c in next(ws_export.iter_rows(min_row=1, max_row=1))]
    assert naglowki_export == eksport.NAGLOWKI

    # typy komórek: kanonicznie czyste, nie odtwarzają niespójności źródła.
    # PNI jest WYJĄTKIEM od 0.1-alpha.3.2 - zostaje TEKSTEM, patrz
    # test_pni_eksportuje_sie_jako_tekst_zachowujac_zera_wiodace
    for row in ws_export.iter_rows(min_row=2):
        assert isinstance(row[0].value, date)          # data
        assert isinstance(row[5].value, int)            # ilosc_total
        if row[7].value is not None:                    # PNI ZPO
            assert isinstance(row[7].value, str)


def test_round_trip_rejon_smieciowy_eksportuje_sie_jako_kanoniczny(conn, tmp_path):
    # realna próbka ma 0 śmieciowych rejonów (patrz plan 0.1-alpha.3.1,
    # sekcja Weryfikacja) - dowodzi tylko braku regresji, nie samej
    # normalizacji. Fiksturka syntetyczna dowodzi, że "???" faktycznie
    # przechodzi przez cały łańcuch import -> zapis -> export.
    wiersz = {
        "data": date(2026, 8, 10), "Kurier": "Kowalski Jan",
        " Pełna Nazwa Nadawcy": "Żabka", "Adres odbioru dla wszystkich nadawców": "Odkryta 24",
        "Rejon": "-", "Wykonawca": "Koli", "PNI ZPO": "228648",
        " Wpisujemy łączną liczbę odebranych Pocztexów": 3,
        " Wpisujemy   w tym liczbę z Zewnetrznych Punktów Odbiorów ": 3,
    }
    wynik = import_row(conn, wiersz)
    assert not wynik["skipped"]

    sciezka = tmp_path / "export.xlsx"
    eksport.eksportuj_miesiac(conn, 2026, 8, sciezka)

    ws = openpyxl.load_workbook(sciezka)[nazwa_arkusza(2026, 8)]
    wiersz_eksportu = next(ws.iter_rows(min_row=2, max_row=2))
    indeks_rejonu = eksport.NAGLOWKI.index("Rejon")
    assert wiersz_eksportu[indeks_rejonu].value == REJON_NIEZNANY


# --- 0.1-alpha.3.2: PNI jako tekst (koercja do int gubiła zera wiodące) ---

def test_pni_eksportuje_sie_jako_tekst_zachowujac_zera_wiodace(conn, tmp_path):
    # do 0.1-alpha.3.1 eksport rzutował PNI "007" -> int 7, a reimport czytał
    # "7" - ten sam fizyczny punkt dostawał DWA różne klucze i powstawał
    # duplikat. Samo-zadana korupcja, bez udziału żadnego obcego pliku.
    kurier_id = conn.execute(
        "INSERT INTO kurierzy (imie_nazwisko) VALUES ('Kowalski Jan')").lastrowid
    punkt_id, _ = get_or_create_punkt(conn, "Żabka", "Odkryta 24", "007")
    conn.execute(
        "INSERT INTO transakcje (data, kurier_id, punkt_id, ilosc_total)"
        " VALUES ('2026-08-03', ?, ?, 3)", (kurier_id, punkt_id))

    sciezka = tmp_path / "export.xlsx"
    eksport.eksportuj_miesiac(conn, 2026, 8, sciezka)

    ws = openpyxl.load_workbook(sciezka)[nazwa_arkusza(2026, 8)]
    wiersz = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    assert wiersz[eksport.NAGLOWKI.index("PNI ZPO")] == "007"


# --- 0.1-alpha.3.2: znacznik pochodzenia + odcisk palca danych ---

def test_eksport_zapisuje_znacznik_i_odcisk(conn, tmp_path):
    _wstaw_prosta_transakcje(conn)
    sciezka = tmp_path / "export.xlsx"
    eksport.eksportuj_miesiac(conn, 2026, 8, sciezka)

    wb = openpyxl.load_workbook(sciezka)
    nazwy = {p.name for p in wb.custom_doc_props.props}
    assert eksport.NAZWA_ZNACZNIKA in nazwy
    assert eksport.NAZWA_ODCISKU in nazwy


def test_wlasny_nietkniety_eksport_jest_zaufany(conn, tmp_path):
    _wstaw_prosta_transakcje(conn)
    sciezka = tmp_path / "export.xlsx"
    eksport.eksportuj_miesiac(conn, 2026, 8, sciezka)

    assert eksport.zweryfikuj_plik(sciezka) == eksport.PLIK_ZAUFANY


def test_plik_bez_znacznika_jest_obcy(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.append(eksport.NAGLOWKI)
    sciezka = tmp_path / "obcy.xlsx"
    wb.save(sciezka)

    assert eksport.zweryfikuj_plik(sciezka) == eksport.PLIK_OBCY


def test_plik_ze_znacznikiem_ale_zmieniona_komorka_jest_zmodyfikowany(conn, tmp_path):
    # sedno decyzji Papavera: ludziom modyfikującym pliki Excela nie ufamy -
    # znacznik bez zgodnego odcisku NIE może uchodzić za zaufany
    _wstaw_prosta_transakcje(conn)
    sciezka = tmp_path / "export.xlsx"
    eksport.eksportuj_miesiac(conn, 2026, 8, sciezka)

    wb = openpyxl.load_workbook(sciezka)
    ws = wb[nazwa_arkusza(2026, 8)]
    ws.cell(row=2, column=6).value = 999  # ktoś "poprawił" ilość w Excelu
    wb.save(sciezka)

    assert eksport.zweryfikuj_plik(sciezka) == eksport.PLIK_ZMODYFIKOWANY


def test_odcisk_zalezy_od_zawartosci_a_nie_od_kolejnosci_zapisu(conn, tmp_path):
    # dwa eksporty tych samych danych muszą dać ten sam odcisk - inaczej
    # własny plik po prostu nigdy by się nie zweryfikował
    _wstaw_prosta_transakcje(conn)
    a, b = tmp_path / "a.xlsx", tmp_path / "b.xlsx"
    eksport.eksportuj_miesiac(conn, 2026, 8, a)
    eksport.eksportuj_miesiac(conn, 2026, 8, b)

    def odcisk(sciezka):
        wb = openpyxl.load_workbook(sciezka)
        return wb.custom_doc_props[eksport.NAZWA_ODCISKU].value

    assert odcisk(a) == odcisk(b)


def test_zweryfikuj_plik_nieczytelny_nie_wybucha(tmp_path):
    sciezka = tmp_path / "uszkodzony.xlsx"
    sciezka.write_text("to nie jest xlsx", encoding="utf-8")
    assert eksport.zweryfikuj_plik(sciezka) == eksport.PLIK_OBCY
