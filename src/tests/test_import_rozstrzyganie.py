"""
Ekran korekty importu (0.1-alpha.4): każda niespójność wymaga decyzji,
albo świadomie pomijasz całość i dostajesz plik do poprawy.

Do 0.1-alpha.4 dało się kliknąć „Zatwierdź import" z nietkniętymi
różnicami w zapisie nazwisk - i wchodziły one do bazy jako osobni
kurierzy. Decyzja zapadała przez NIEKLIKANIE, czyli najgorszym możliwym
sposobem: bez śladu, bez świadomości i bez możliwości odróżnienia
„zdecydowałem, że to dwie różne osoby" od „nie zauważyłem".

Wymaga środowiska graficznego - pomijany tak samo jak reszta testów GUI.
"""
import subprocess
import sys
import tkinter as tk
from datetime import date

import openpyxl
import pytest

from zpo_tracker import eksport, repo
from zpo_tracker.gui import zakladka_import_export as zie
from zpo_tracker.gui.zakladka_import_export import DialogKorektyImportu
from zpo_tracker.import_orchestrator import KLUCZ_NUMERU_WIERSZA, zwaliduj_wiersze
from zpo_tracker.normalizacja import Podobienstwo


def _ma_display():
    try:
        wynik = subprocess.run(
            [sys.executable, "-c",
             "import tkinter as tk; r = tk.Tk(); tk.Entry(r).pack(); r.update(); r.destroy()"],
            capture_output=True, timeout=10,
        )
        return wynik.returncode == 0
    except Exception:
        return False


pytestmark = pytest.mark.skipif(
    not _ma_display(), reason="wymaga środowiska graficznego (DISPLAY)"
)


@pytest.fixture
def root():
    r = tk.Tk()
    yield r
    r.destroy()


@pytest.fixture
def conn():
    c = repo.polacz(":memory:")
    repo.utworz_schemat(c)
    yield c
    c.close()


def _surowy(numer=2, **nadpisz):
    dane = {
        "data": date(2026, 8, 3),
        " Pełna Nazwa Nadawcy": "Żabka",
        "Adres odbioru dla wszystkich nadawców": "Odkryta 24",
        "Kurier": "Kowalski Jan",
        "Rejon": "WA87",
        " Wpisujemy łączną liczbę odebranych Pocztexów": 3,
        "PNI ZPO": "228648",
        KLUCZ_NUMERU_WIERSZA: numer,
    }
    dane.update(nadpisz)
    return dane


def _zbuduj(root, conn, tmp_path, ostrzezenia=(), surowe=None, odrzucone=None):
    surowe = surowe if surowe is not None else [_surowy()]
    zwalidowane, auto_odrzucone = zwaliduj_wiersze(surowe)
    zrodlo = tmp_path / "sierpien.xlsx"
    zrodlo.write_bytes(b"")  # sam plik nie jest czytany, liczy się ścieżka
    return DialogKorektyImportu(
        root, conn, str(tmp_path), "sierpien.xlsx", zwalidowane,
        odrzucone if odrzucone is not None else auto_odrzucone,
        propozycje=[], ostrzezenia=list(ostrzezenia),
        on_gotowe=lambda w: None,
        sciezka_pliku=zrodlo, surowe=surowe,
    )


# --- wymuszenie decyzji -------------------------------------------------

def test_nietkniete_roznice_sa_nierozstrzygniete(root, conn, tmp_path):
    d = _zbuduj(root, conn, tmp_path,
                ostrzezenia=[Podobienstwo("Wołczuk Rafał", "Wolczuk Rafal")])
    assert len(d.nierozstrzygniete()) == 1
    d.destroy()


def test_wybor_formy_rozstrzyga(root, conn, tmp_path):
    o = Podobienstwo("Wołczuk Rafał", "Wolczuk Rafal")
    d = _zbuduj(root, conn, tmp_path, ostrzezenia=[o])
    d.rozstrzygniecia[(o.a, o.b)] = o.a
    assert d.nierozstrzygniete() == []
    d.destroy()


def test_zostaw_obie_tez_jest_decyzja(root, conn, tmp_path):
    """Sedno zmiany: „to naprawdę dwie różne osoby" musi dać się odróżnić
    od „nie zauważyłem"."""
    o = Podobienstwo("Wołczuk Rafał", "Wolczuk Rafal")
    d = _zbuduj(root, conn, tmp_path, ostrzezenia=[o])
    d.rozstrzygniecia[(o.a, o.b)] = None
    assert d.nierozstrzygniete() == []
    assert o.b not in d.mapowanie_z_ostrzezen
    d.destroy()


def test_zatwierdz_nie_importuje_przy_nierozstrzygnietych(root, conn, tmp_path, monkeypatch):
    wolane = []
    monkeypatch.setattr(zie.messagebox, "showwarning",
                        lambda *a, **k: wolane.append("ostrzezenie"))
    d = _zbuduj(root, conn, tmp_path,
                ostrzezenia=[Podobienstwo("Wołczuk Rafał", "Wolczuk Rafal")])
    monkeypatch.setattr(d, "_wykonaj_import", lambda: wolane.append("import"))
    d._zatwierdz()
    assert wolane == ["ostrzezenie"]
    d.destroy()


def test_zatwierdz_importuje_gdy_wszystko_rozstrzygniete(root, conn, tmp_path, monkeypatch):
    o = Podobienstwo("Wołczuk Rafał", "Wolczuk Rafal")
    d = _zbuduj(root, conn, tmp_path, ostrzezenia=[o])
    d.rozstrzygniecia[(o.a, o.b)] = o.a
    wolane = []
    monkeypatch.setattr(d, "_wykonaj_import", lambda: wolane.append("import"))
    d._zatwierdz()
    assert wolane == ["import"]
    d.destroy()


# --- świadome pominięcie ------------------------------------------------

def test_pominiecie_wymaga_potwierdzenia(root, conn, tmp_path, monkeypatch):
    """Odmowa w oknie potwierdzenia ma nie importować niczego."""
    monkeypatch.setattr(zie.messagebox, "askyesno", lambda *a, **k: False)
    d = _zbuduj(root, conn, tmp_path,
                ostrzezenia=[Podobienstwo("Wołczuk Rafał", "Wolczuk Rafal")])
    wolane = []
    monkeypatch.setattr(d, "_wykonaj_import", lambda: wolane.append("import"))
    d._pomin_niespojnosci()
    assert wolane == []
    d.destroy()


def test_pominiecie_po_potwierdzeniu_importuje(root, conn, tmp_path, monkeypatch):
    monkeypatch.setattr(zie.messagebox, "askyesno", lambda *a, **k: True)
    d = _zbuduj(root, conn, tmp_path,
                ostrzezenia=[Podobienstwo("Wołczuk Rafał", "Wolczuk Rafal")])
    wolane = []
    monkeypatch.setattr(d, "_wykonaj_import", lambda: wolane.append("import"))
    d._pomin_niespojnosci()
    assert wolane == ["import"]
    d.destroy()


# --- pliki do poprawy ---------------------------------------------------

def test_bez_odrzuconych_nie_powstaja_pliki(root, conn, tmp_path):
    """Nie zaśmiecamy katalogu użytkownika plikiem, gdy wszystko weszło."""
    d = _zbuduj(root, conn, tmp_path)
    assert d._zapisz_do_poprawy({"wymagajace_uwagi": []}) == []
    d.destroy()


def test_pliki_ladu_obok_zrodla(root, conn, tmp_path):
    """Obok pliku źródłowego, nie w katalogu danych aplikacji - użytkownik
    wie, gdzie położył swój Excel, a %LOCALAPPDATA% jest dla niego
    miejscem, którego nie znajdzie."""
    surowe = [_surowy(numer=2), _surowy(numer=3, **{"Kurier": "Nowak Anna"})]
    d = _zbuduj(root, conn, tmp_path, surowe=surowe,
                odrzucone=[{"wiersz": surowe[1], "powod": "zła ilość"}])
    pliki = d._zapisz_do_poprawy({"wymagajace_uwagi": []})
    assert [p.name for p in pliki] == [
        "sierpien-do-poprawy.xlsx", "sierpien-odrzucone.xlsx"]
    assert all(p.parent == tmp_path for p in pliki)
    d.destroy()


def test_plik_do_poprawy_ma_tylko_niezaimportowane(root, conn, tmp_path):
    surowe = [_surowy(numer=2), _surowy(numer=3, **{"Kurier": "Nowak Anna"})]
    d = _zbuduj(root, conn, tmp_path, surowe=surowe,
                odrzucone=[{"wiersz": surowe[1], "powod": "zła ilość"}])
    reszta, _ = d._zapisz_do_poprawy({"wymagajace_uwagi": []})

    ws = openpyxl.load_workbook(reszta).active
    naglowki = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    wiersze = [dict(zip(naglowki, w)) for w in ws.iter_rows(min_row=2, values_only=True)]
    assert len(wiersze) == 1
    assert wiersze[0]["Kurier"] == "Nowak Anna"
    assert wiersze[0][eksport.NAGLOWEK_POWODU] == "zła ilość"
    assert KLUCZ_NUMERU_WIERSZA not in naglowki
    d.destroy()
