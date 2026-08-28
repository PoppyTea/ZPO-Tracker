"""
Porządkowanie powtórzeń w arkuszu PRZED importem.

Narzędzie na czas przejściowy. Miesiące zaległości są sklejone z pracy
kilku osób i ten sam odbiór bywa wpisany dwa razy — dopóki nie zrobi tego
auto-naprawa w potoku importu, robi to jedno polecenie na pliku:

    python -m zpo_tracker.porzadkowanie sierpien.xlsx

Powstaje `sierpien-POSPRZATANY.xlsx` obok źródła. **Źródło zostaje
nietknięte** — narzędzie działające na miesiącu rozliczeń nie ma prawa
nadpisywać jedynej kopii.

REGUŁA wzięta z pomiaru na realnym sierpniu, nie z wyobraźni. Odległości
między powtórzeniami układały się tam tak:

    1,1,1,1,1,1,1,1,2,2,2,3,4,4,4,4,7,8,9 | 11,12,31,47,51,53,80,245,247,3429,3431

Przerwa między 9 a 11 jest naturalną granicą i stąd próg dziesięciu
wierszy:

* **blisko** (do 10 wierszy) — jeden blok arkusza, ktoś rozbił jeden
  odbiór na dwa wiersze. Scalamy, **ilości sumujemy**.
* **daleko** — dwie osoby wpisujące w dwóch częściach sklejonego pliku.
  Nie wiadomo, czy to dwa odbiory, czy dwa opisy jednego, więc automat
  **nie rozstrzyga**: oba wiersze zostają, oznaczone do decyzji.

Nazw firm narzędzie NIE RUSZA. Dwie pisownie to dla niego dwa różne
punkty — scalanie nazw jest osobną decyzją (aliasy) i osobną robotą.
"""
import argparse
import collections
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

# Próg "blisko/daleko" w wierszach - patrz docstring modułu.
PROG_BLISKOSCI = 10

NAGLOWEK_UWAGI = "UWAGA (dodane przez porządkowanie)"

# Klucz powtórzenia. Świadomie BEZ normalizacji nazw: dwie pisownie tej
# samej firmy zostają dwoma punktami, bo ich scalanie to osobna decyzja
# człowieka, a nie efekt uboczny sprzątania duplikatów.
POLA_KLUCZA = ("data", "Kurier", " Pełna Nazwa Nadawcy",
               "Adres odbioru dla wszystkich nadawców")

_TLO_SCALONE = PatternFill("solid", fgColor="FFF3CD")
_TLO_DECYZJA = PatternFill("solid", fgColor="F8D7DA")


@dataclass
class Raport:
    wierszy_przed: int = 0
    wierszy_po: int = 0
    scalone: int = 0
    do_decyzji: int = 0
    sumy_zgodne: bool = True
    sumy: dict = field(default_factory=dict)

    def opis(self) -> str:
        linie = [
            f"wierszy przed:  {self.wierszy_przed}",
            f"wierszy po:     {self.wierszy_po}   "
            f"(usunięto {self.wierszy_przed - self.wierszy_po})",
            f"scalonych grup: {self.scalone}   (ilości zsumowane)",
            f"do decyzji:     {self.do_decyzji} grup, "
            f"{self.do_decyzji * 2} wierszy na czerwono",
            "",
            "KONTROLA SUM (muszą się zgadzać co do sztuki):",
        ]
        for nazwa, (przed, po) in self.sumy.items():
            znak = "OK  " if przed == po else "BŁĄD"
            linie.append(f"  {znak}  {nazwa.strip()[:48]:50} {przed} -> {po}")
        if not self.sumy_zgodne:
            linie.append("")
            linie.append("!!! SUMY SIĘ NIE ZGADZAJĄ - NIE UŻYWAJ TEGO PLIKU.")
        return "\n".join(linie)


def _czy_ilosc(naglowek) -> bool:
    return bool(naglowek) and (
        "Wpisujemy" in naglowek or "w tym" in naglowek or "nierozlicz" in naglowek)


def _sumy(wiersze, naglowki) -> dict:
    """Sumy wszystkich kolumn liczbowych. Osobno liczone dla źródła
    i wyniku - to jedyny niezmiennik, który MUSI trzymać."""
    out = {}
    for i, n in enumerate(naglowki):
        if not _czy_ilosc(n):
            continue
        out[n] = sum(w[i] for w in wiersze
                     if i < len(w) and isinstance(w[i], (int, float)))
    return out


def porzadkuj(zrodlo, wynik, prog=PROG_BLISKOSCI, arkusz=None) -> Raport:
    """
    Czyta `zrodlo`, zapisuje uporządkowaną kopię do `wynik`, zwraca `Raport`.

    Źródło nie jest modyfikowane. Kontrola sum liczona jest z DWÓCH
    NIEZALEŻNYCH ODCZYTÓW list, nie z jednej mutowanej w miejscu — na tym
    właśnie błędzie dałem się złapać przy pierwszym przebiegu i wynik
    wyglądał na rozjechany, choć dane były poprawne.
    """
    wb = openpyxl.load_workbook(zrodlo, data_only=True)
    ws = wb[arkusz] if arkusz else wb[wb.sheetnames[0]]
    naglowki = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    indeks = {n: i for i, n in enumerate(naglowki) if n}
    oryginalne = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]
    sumy_przed = _sumy(oryginalne, naglowki)

    dane = [list(r) for r in oryginalne]      # kopia do modyfikacji
    brakujace = [p for p in POLA_KLUCZA if p not in indeks]
    if brakujace:
        raise ValueError(
            "Arkusz nie ma kolumn potrzebnych do rozpoznania powtórzeń: "
            + ", ".join(brakujace))

    klucz = lambda w: tuple(str(w[indeks[p]] or "").strip() for p in POLA_KLUCZA)
    grupy = collections.defaultdict(list)
    for i, w in enumerate(dane):
        grupy[klucz(w)].append(i)

    raport = Raport(wierszy_przed=len(dane))
    do_usuniecia, uwagi = set(), {}

    for pozycje in grupy.values():
        if len(pozycje) < 2:
            continue
        odleglosc = pozycje[-1] - pozycje[0]
        if odleglosc <= prog:
            _scal(dane, pozycje, naglowki, indeks)
            do_usuniecia.update(pozycje[1:])
            uwagi[pozycje[0]] = (
                f"SCALONO {len(pozycje)} wiersze (odległość {odleglosc}), "
                f"ilości zsumowane")
            raport.scalone += 1
        else:
            for p in pozycje:
                inne = ", ".join(str(q + 2) for q in pozycje if q != p)
                uwagi[p] = (f"DO DECYZJI: powtórzenie wiersza {inne} "
                            f"(odległość {odleglosc})")
            raport.do_decyzji += 1

    koncowe = [w for i, w in enumerate(dane) if i not in do_usuniecia]
    raport.wierszy_po = len(koncowe)

    sumy_po = _sumy(koncowe, naglowki)
    raport.sumy = {n: (sumy_przed[n], sumy_po.get(n, 0)) for n in sumy_przed}
    raport.sumy_zgodne = all(a == b for a, b in raport.sumy.values())

    _zapisz(wynik, ws.title, naglowki, koncowe,
            [uwagi.get(i, "") for i in range(len(dane)) if i not in do_usuniecia])
    return raport


def _scal(dane, pozycje, naglowki, indeks):
    """
    Scala grupę w pierwszy wiersz: ilości sumuje, pola nieliczbowe
    uzupełnia PIERWSZĄ NIEPUSTĄ wartością z grupy.

    To drugie jest istotne: PNI albo rejon bywa wpisany tylko w jednym
    z dwóch wierszy, a scalanie, które by go zgubiło, zamieniałoby
    porządkowanie w utratę danych.
    """
    baza = dane[pozycje[0]]
    for nazwa, i in indeks.items():
        if _czy_ilosc(nazwa):
            wartosci = [dane[p][i] for p in pozycje
                        if i < len(dane[p]) and isinstance(dane[p][i], (int, float))]
            if wartosci:
                baza[i] = sum(wartosci)
        elif not str(baza[i] or "").strip():
            for p in pozycje[1:]:
                if i < len(dane[p]) and str(dane[p][i] or "").strip():
                    baza[i] = dane[p][i]
                    break


def _zapisz(sciezka, tytul, naglowki, wiersze, uwagi):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tytul
    ws.append(list(naglowki) + [NAGLOWEK_UWAGI])
    ws.cell(1, len(naglowki) + 1).font = Font(bold=True)

    for wiersz, uwaga in zip(wiersze, uwagi):
        ws.append(list(wiersz) + [uwaga])
        if uwaga.startswith("SCALONO"):
            ws.cell(ws.max_row, len(naglowki) + 1).fill = _TLO_SCALONE
        elif uwaga.startswith("DO DECYZJI"):
            # CAŁY wiersz na czerwono, nie sama uwaga: te wymagają
            # przeczytania w całości, żeby porównać je z parą.
            for kolumna in range(1, len(naglowki) + 2):
                ws.cell(ws.max_row, kolumna).fill = _TLO_DECYZJA
    wb.save(sciezka)


def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Porządkuje powtórzenia w arkuszu przed importem. "
                    "Źródło zostaje nietknięte.")
    parser.add_argument("zrodlo", help="plik .xlsx do uporządkowania")
    parser.add_argument("-o", "--wynik", help="plik wynikowy "
                        "(domyślnie <źródło>-POSPRZATANY.xlsx obok źródła)")
    parser.add_argument("--prog", type=int, default=PROG_BLISKOSCI,
                        help=f"granica blisko/daleko w wierszach "
                             f"(domyślnie {PROG_BLISKOSCI})")
    parser.add_argument("--arkusz", help="nazwa arkusza (domyślnie pierwszy)")
    args = parser.parse_args(argv)

    zrodlo = Path(args.zrodlo)
    wynik = Path(args.wynik) if args.wynik else \
        zrodlo.with_name(f"{zrodlo.stem}-POSPRZATANY.xlsx")

    raport = porzadkuj(zrodlo, wynik, prog=args.prog, arkusz=args.arkusz)
    print(raport.opis())
    print(f"\nzapisano: {wynik}")
    # Kod wyjścia różny od zera, gdy sumy się nie zgadzają - żeby
    # nieudane porządkowanie dało się wykryć bez czytania wydruku.
    return 0 if raport.sumy_zgodne else 1


if __name__ == "__main__":
    raise SystemExit(main())
