"""
Zakładka import/export. Import: wczytanie .xlsx, ekran korekty pokazujący
WYŁĄCZNIE wiersze wymagające uwagi (odrzucone, propozycje scalenia
literówek do zatwierdzenia/odznaczenia, ostrzeżenia o różnicach
diakrytyków), reszta importuje się cicho z podsumowaniem liczbowym.
Export: wybór miesiąca, układ kolumn identyczny ze snapshotem.

Cała logika jest w import_orchestrator.py/eksport.py - tu tylko zbieranie
wartości z pól, wywołanie i wyświetlenie wyniku.
"""
from datetime import date
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import openpyxl

from zpo_tracker import arkusze, eksport, operacje, rejonarz, ustawienia
from zpo_tracker.gui.roznice import segmenty_roznicy
from zpo_tracker.import_orchestrator import (
    KLUCZ_NUMERU_WIERSZA,
    wybierz_niezaimportowane,
    zbierz_odrzucone,
    zaimportuj,
    znajdz_ostrzezenia_podobienstwa_kurierow,
    znajdz_propozycje_scalenia_kurierow,
    zwaliduj_wiersze,
)


def _pole_z_roznicami(parent, segmenty):
    """Text tylko-do-odczytu z fragmentami różniącymi się pogrubionym,
    kolorowym tekstem (GH #2) - segmenty z gui.roznice.segmenty_roznicy."""
    szerokosc = max(sum(len(t) for t, _ in segmenty), 4)
    tlo = ttk.Style().lookup("TFrame", "background") or "SystemButtonFace"
    pole = tk.Text(
        parent, height=1, width=szerokosc, wrap="none", borderwidth=0,
        highlightthickness=0, background=tlo,
    )
    pole.tag_configure("rozni", foreground="#c0392b", font=("TkDefaultFont", 10, "bold"))
    for tekst, rozni in segmenty:
        pole.insert("end", tekst, "rozni" if rozni else ())
    pole.configure(state="disabled")
    return pole


# Nazwy pól z modelu na nagłówki, które użytkownik widzi w swoim Excelu.
# Bez tego raport mówiłby "ilosc_zpo", czego nie ma w żadnej kolumnie
# jego pliku, więc nie miałby jak tego odnaleźć.
_ETYKIETY_POL = {
    p: n for n, p in
    __import__("zpo_tracker.import_orchestrator", fromlist=["MAPA_NAGLOWKOW"])
    .MAPA_NAGLOWKOW.items()
}


def _po_polsku(istniejace):
    return {_ETYKIETY_POL.get(k, k): v for k, v in istniejace.items()}


def _podsumowanie_wczytania(wczytane):
    """
    Zdanie o tym, co weszło z eksportu BaŚKi.

    Liczby POMINIĘTYCH są tu tak samo ważne jak wczytanych: użytkownik
    wskazuje plik ogólnopolski i musi zobaczyć, że z 22 tysięcy wierszy
    wzięliśmy dwa i pół - inaczej wygląda to na awarię importu, a jest
    poprawnym filtrowaniem.
    """
    w = wczytane.wynik
    if wczytane.rodzaj == rejonarz.RODZAJ_PUNKTY_ZPO:
        czesci = [f"Punkty ZPO: wczytano {w.zapisane}"]
        if w.pominiete:
            czesci.append(f"pominięto {w.pominiete} punktów spoza naszych "
                          f"placówek ({', '.join(sorted(rejonarz.WEZLY_PUNKTOW))}) "
                          f"- tak ma być, plik jest ogólnopolski")
        if w.bez_pni:
            czesci.append(f"{w.bez_pni} bez PNI")
        if w.bez_filtrowania:
            czesci.append("UWAGA: ten eksport nie mówi, z której placówki są "
                          "punkty, więc wzięto CAŁY plik ogólnopolski")
        return ", ".join(czesci) + "."

    czesci = [f"Rejonarz: wczytano {w.zapisane} adresów"]
    if w.pominiete:
        czesci.append(f"pominięto {w.pominiete} adresów spoza naszej "
                      f"placówki ({rejonarz.WEZEL_ZPO}) - tak ma być")
    if w.bez_rejonu:
        czesci.append(f"{w.bez_rejonu} bez rejonu")
    if w.arkusze_pominiete:
        czesci.append(f"pominięto {len(w.arkusze_pominiete)} arkuszy: "
                      + ", ".join(w.arkusze_pominiete[:5])
                      + ("..." if len(w.arkusze_pominiete) > 5 else ""))
    if w.rejon_z_nazw_arkuszy:
        # Ten eksport z definicji nie ma kolumny węzła - straszenie jej
        # brakiem byłoby myleniem normalnego kształtu pliku z usterką.
        # Ryzyko jest inne i konkretne: to może być eksport CUDZEGO węzła.
        czesci.append(f"rejon wzięty z nazw arkuszy - upewnij się, że to "
                      f"eksport NASZEJ placówki ({rejonarz.WEZEL_ZPO}), "
                      f"bo plik sam tego nie mówi")
    elif w.bez_filtrowania:
        # Nie "brak kolumn Węzeł/TK" - to nazywa PRZYCZYNĘ techniczną
        # i nie mówi nic o tym, co z tym zrobić. Użytkownik ma wiedzieć,
        # co mu grozi: podpowiedziany rejon z cudzej placówki.
        czesci.append("UWAGA: ten eksport nie mówi, z której placówki są "
                      "adresy, więc wzięto wszystkie - jeśli obejmuje inne "
                      "WER-y, w słowniku znajdą się cudze rejony")
    return ", ".join(czesci) + "."


def _wczytaj_surowe_wiersze(sciezka):
    wb = openpyxl.load_workbook(sciezka, data_only=True)
    ws = wb[wb.sheetnames[0]]
    naglowki = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    # Numer wiersza doklejany od razu przy czytaniu - bez niego raport
    # odrzuconych mówi "71 wierszy wymagało uwagi" i nie da się z tym nic
    # zrobić. `_przemapuj` w orchestratorze filtruje po MAPA_NAGLOWKOW,
    # więc ta metadana nigdy nie dociera do WierszImportu.
    return [
        dict(zip(naglowki, wiersz),
             **{KLUCZ_NUMERU_WIERSZA: numer})
        for numer, wiersz in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2)
    ]


class DialogKorektyImportu(tk.Toplevel):
    """Ekran korekty: tylko to, co wymaga uwagi. Reszta importuje się cicho."""

    def __init__(self, parent, conn, katalog_danych, nazwa_pliku, zwalidowane,
                 odrzucone, propozycje, ostrzezenia, on_gotowe,
                 autor_id=None, sesja_uuid=None, status_zaufania=eksport.PLIK_OBCY,
                 sciezka_pliku=None, surowe=None, szukaj_rejonarza=None):
        super().__init__(parent)
        self.title("Korekta importu")
        self.geometry("640x560")
        self.conn = conn
        self.katalog_danych = katalog_danych
        self.nazwa_pliku = nazwa_pliku
        self.zwalidowane = zwalidowane
        self.odrzucone = odrzucone
        self.ostrzezenia = ostrzezenia
        self.on_gotowe = on_gotowe
        self.szukaj_rejonarza = szukaj_rejonarza
        # 0.1-alpha.3.2: atrybucja i sesja dla wierszy z importu - dotąd
        # import nie pisał żadnej z nich, patrz import_orchestrator.zaimportuj.
        self.autor_id = autor_id
        self.sesja_uuid = sesja_uuid
        self.status_zaufania = status_zaufania
        self.var_wymus_zaufanie = tk.BooleanVar(value=False)
        self.checkbox_wymuszenia = None
        self.zmienne_propozycji = []
        self.mapowanie_z_ostrzezen = {}
        self.sciezka_pliku = Path(sciezka_pliku) if sciezka_pliku else None
        self.surowe = surowe or []
        # Rozstrzygnięcia różnic w zapisie. Klucz to para nazw, wartość -
        # jawna decyzja użytkownika. "Zostaw obie" MUSI być osobnym
        # wpisem, nie brakiem wpisu: inaczej "jeszcze się nie zdecydowałem"
        # i "świadomie zostawiam obie" wyglądają identycznie, a to dwie
        # zupełnie różne sytuacje.
        self.rozstrzygniecia = {}

        ttk.Label(
            self,
            text=f"{len(zwalidowane)} wierszy do zaimportowania. "
                 f"Poniżej tylko to, co wymaga uwagi - reszta wejdzie cicho.",
            wraplength=600, justify="left",
        ).pack(anchor="w", padx=10, pady=(10, 6))

        self._zbuduj_panel_zaufania()

        notebook = ttk.Notebook(self)
        notebook.pack(fill="both", expand=True, padx=10, pady=6)

        if odrzucone:
            ramka = ttk.Frame(notebook)
            notebook.add(ramka, text=f"Odrzucone ({len(odrzucone)})")
            tekst = tk.Text(ramka, wrap="word")
            tekst.pack(fill="both", expand=True)
            for o in odrzucone:
                tekst.insert("end", f"• {o['wiersz'].get('Kurier', '?')}: {o['powod']}\n")
            tekst.configure(state="disabled")

        ramka_literowki = ttk.Frame(notebook)
        notebook.add(ramka_literowki, text=f"Prawdopodobne literówki ({len(propozycje)})")
        if not propozycje:
            ttk.Label(ramka_literowki, text="Brak.").pack(anchor="w", padx=6, pady=6)
        for p in propozycje:
            var = tk.BooleanVar(value=True)
            ttk.Checkbutton(
                ramka_literowki, variable=var,
                text=f"Scalić „{p['z']}” → „{p['na']}”",
            ).pack(anchor="w", padx=6, pady=2)
            self.zmienne_propozycji.append((p, var))

        if ostrzezenia:
            ramka_o = ttk.Frame(notebook)
            notebook.add(ramka_o, text=f"Różnice w zapisie ({len(ostrzezenia)})")
            ttk.Label(
                ramka_o,
                text="Różnią się tylko wielkością liter/polskimi znakami - NIE scalono "
                     "automatycznie. Wybierz, która forma ma zostać (druga zostanie do "
                     "niej scalona), albo kliknij „obie”, jeśli to naprawdę dwie różne "
                     "osoby. Każda pozycja wymaga decyzji - „obie” też jest decyzją.",
                wraplength=580, justify="left",
            ).pack(anchor="w", padx=6, pady=6)
            for o in ostrzezenia:
                self._dodaj_wiersz_ostrzezenia(ramka_o, o)

        # Przełącznik pliku ze znalezionymi duplikatami. Stan wraca
        # z `settings.json` i tam wraca po zmianie - odznaczanie go przy
        # każdym imporcie byłoby karą za jednorazową decyzję.
        self.var_zapisz_duplikaty = tk.BooleanVar(
            value=ustawienia.czy_zapisywac_duplikaty(
                ustawienia.wczytaj(self.katalog_danych)))
        ttk.Checkbutton(
            self, variable=self.var_zapisz_duplikaty,
            text="Zapisz plik ze znalezionymi duplikatami (obok pliku źródłowego)",
        ).pack(anchor="w", padx=10)

        pasek = ttk.Frame(self)
        pasek.pack(fill="x", padx=10, pady=10)
        ttk.Button(pasek, text="Zatwierdź import", command=self._zatwierdz).pack(side="right")
        ttk.Button(
            pasek, text="Pomiń niespójności", command=self._pomin_niespojnosci,
        ).pack(side="right", padx=6)
        ttk.Button(pasek, text="Anuluj", command=self.destroy).pack(side="right", padx=6)

    def _zbuduj_panel_zaufania(self):
        """
        Informacja o pochodzeniu pliku + (warunkowo) przełącznik wymuszenia.

        Przełącznik pojawia się WYŁĄCZNIE gdy: (a) settings.json ma wpis
        odsłaniający tryb zaawansowany ORAZ (b) plik jest po prostu obcy,
        nie sfałszowany. Plik z NASZYM znacznikiem, ale niezgodnym odciskiem,
        nie może zostać uznany za zaufany żadną drogą - to nie jest "obcy
        plik, któremu ktoś świadomie ufa", tylko nasz plik zmodyfikowany po
        wyjściu z programu (decyzja Papavera, patrz eksport.zweryfikuj_plik).
        """
        ramka = ttk.Frame(self)
        ramka.pack(fill="x", padx=10, pady=(0, 6))

        if self.status_zaufania == eksport.PLIK_ZAUFANY:
            ttk.Label(
                ramka, text="✔ To plik wyeksportowany z tego programu i niezmieniony "
                            "- PNI i rejony zostaną zaimportowane.",
                foreground="#1e7a3a", wraplength=600, justify="left",
            ).pack(anchor="w")
            return

        if self.status_zaufania == eksport.PLIK_ZMODYFIKOWANY:
            komunikat = ("⚠ Ten plik pochodzi z tego programu, ale jego zawartość "
                          "została zmieniona poza nim. PNI i rejony NIE zostaną "
                          "zaimportowane i nie da się tego pominąć.")
        else:
            komunikat = ("⚠ Plik spoza tego programu. PNI i rejony NIE zostaną "
                          "zaimportowane - reszta danych wejdzie normalnie.")
        ttk.Label(
            ramka, text=komunikat, foreground="#c0392b",
            wraplength=600, justify="left",
        ).pack(anchor="w")

        zaawansowane = ustawienia.wczytaj(self.katalog_danych).get("zaawansowane")
        if not isinstance(zaawansowane, dict):
            zaawansowane = {}
        if (self.status_zaufania == eksport.PLIK_OBCY
                and zaawansowane.get("pokaz_wymuszenie_zaufania")):
            self.checkbox_wymuszenia = ttk.Checkbutton(
                ramka, text="Mimo to potraktuj ten plik jako zaufany (zaawansowane)",
                variable=self.var_wymus_zaufanie,
            )
            self.checkbox_wymuszenia.pack(anchor="w", pady=(4, 0))

    def czy_zaufany(self):
        """Czy import ma wnieść PNI i rejon - patrz `_zbuduj_panel_zaufania`."""
        if self.status_zaufania == eksport.PLIK_ZAUFANY:
            return True
        if self.status_zaufania == eksport.PLIK_ZMODYFIKOWANY:
            return False  # nigdy, niezależnie od przełącznika
        return bool(self.var_wymus_zaufanie.get())

    def _dodaj_wiersz_ostrzezenia(self, parent, ostrzezenie):
        """Jeden wiersz konfliktu (GH #2 podświetlenie różnic + GH #1 wybór
        zwycięzcy kliknięciem, zamiast zmuszania do wyjścia do Słowników)."""
        wiersz = ttk.Frame(parent)
        wiersz.pack(fill="x", padx=6, pady=3)

        seg_a, seg_b = segmenty_roznicy(ostrzezenie.a, ostrzezenie.b)
        etykieta_wyniku = ttk.Label(wiersz, text="", foreground="#1e7a3a")

        klucz = (ostrzezenie.a, ostrzezenie.b)

        def wybierz(kanoniczny, odrzucony):
            self.mapowanie_z_ostrzezen[odrzucony] = kanoniczny
            self.mapowanie_z_ostrzezen.pop(kanoniczny, None)
            self.rozstrzygniecia[klucz] = kanoniczny
            etykieta_wyniku.configure(text=f"→ zostaje „{kanoniczny}”")

        def zostaw_obie():
            self.mapowanie_z_ostrzezen.pop(ostrzezenie.a, None)
            self.mapowanie_z_ostrzezen.pop(ostrzezenie.b, None)
            self.rozstrzygniecia[klucz] = None
            etykieta_wyniku.configure(text="→ zostają obie")

        ttk.Button(
            wiersz, text="◄ zostaw", width=8,
            command=lambda: wybierz(ostrzezenie.a, ostrzezenie.b),
        ).pack(side="left")
        _pole_z_roznicami(wiersz, seg_a).pack(side="left", padx=(4, 10))
        _pole_z_roznicami(wiersz, seg_b).pack(side="left", padx=(0, 4))
        ttk.Button(
            wiersz, text="zostaw ►", width=8,
            command=lambda: wybierz(ostrzezenie.b, ostrzezenie.a),
        ).pack(side="left")
        ttk.Button(
            wiersz, text="obie", width=6, command=zostaw_obie,
        ).pack(side="left", padx=(6, 0))
        etykieta_wyniku.pack(side="left", padx=10)

    def nierozstrzygniete(self):
        """Różnice w zapisie, przy których użytkownik jeszcze nic nie
        kliknął. „Zostaw obie” liczy się jako rozstrzygnięcie."""
        return [o for o in self.ostrzezenia
                if (o.a, o.b) not in self.rozstrzygniecia]

    def _zatwierdz(self):
        """Ścieżka pełna: nie da się jej domknąć z nierozstrzygniętą pozycją.

        Do 0.1-alpha.4 dało się kliknąć „Zatwierdź” z nietkniętymi
        różnicami i wchodziły one do bazy jako osobni kurierzy - decyzja
        zapadała przez nieklikanie, czyli najgorszym możliwym sposobem.
        """
        brakujace = self.nierozstrzygniete()
        if brakujace:
            messagebox.showwarning(
                "Nierozstrzygnięte różnice",
                f"Zostało {len(brakujace)} różnic bez decyzji.\n\n"
                "Przy każdej wybierz, która forma ma zostać, albo kliknij "
                "„obie”, jeśli to naprawdę dwie różne osoby.\n\n"
                "Jeśli nie chcesz teraz tego rozstrzygać, użyj przycisku "
                "„Pomiń niespójności”.",
                parent=self,
            )
            return
        self._wykonaj_import()

    def _pomin_niespojnosci(self):
        """Ścieżka druga, jawna: importuj mimo nierozstrzygniętych różnic.

        Osobny przycisk, nie furtka w tamtej - pominięcie ma być
        widoczną decyzją, a nie skutkiem zniecierpliwienia. Potwierdzenie
        mówi wprost, ILE zostanie pominięte i że powstanie plik do poprawy.
        """
        brakujace = self.nierozstrzygniete()
        odpowiedz = messagebox.askyesno(
            "Pominąć niespójności?",
            f"Nierozstrzygniętych różnic w zapisie: {len(brakujace)}.\n"
            "Zostaną zaimportowane jako osobni kurierzy - do scalenia "
            "później w Słownikach.\n\n"
            "Wiersze, które nie wejdą do bazy (odrzucone przy walidacji, "
            "duplikaty, konflikty PNI), zapiszę obok pliku źródłowego "
            "jako osobny plik do poprawy.\n\nKontynuować?",
            parent=self,
        )
        if odpowiedz:
            self._wykonaj_import()

    def _wykonaj_import(self):
        mapowanie = {p["z"]: p["na"] for p, var in self.zmienne_propozycji if var.get()}
        mapowanie.update(self.mapowanie_z_ostrzezen)
        wynik = operacje.wykonaj(
            self.conn, self.katalog_danych, rodzaj="import",
            etykieta=self.nazwa_pliku,
            funkcja=zaimportuj, args=(self.zwalidowane,),
            kwargs={"mapowanie_scalen": mapowanie, "zaufany": self.czy_zaufany(),
                    "autor_id": self.autor_id, "sesja_uuid": self.sesja_uuid,
                    # Bez tego kaskada dedukcji miejscowości nie ma migawki
                    # i zachowuje się tak, jakby jej nie było - a to ona
                    # domyka 75% adresów zamiast 19%.
                    "szukaj": self.szukaj_rejonarza},
            licz_wiersze=lambda w: w["zaimportowano"],
        )
        wynik["pliki"] = self._zapisz_do_poprawy(wynik)
        self.destroy()
        self.on_gotowe(wynik)

    def _zapisz_duplikaty(self, wynik, katalog, rdzen, naglowki):
        """
        Plik ze znalezionymi duplikatami — osobny od pliku-reszty.

        Stan przełącznika zapisuje się do `settings.json` PRZY KAŻDYM
        imporcie, nie tylko przy zmianie: użytkownik odznacza raz i ma
        prawo oczekiwać, że tak zostanie.

        Nazwa mówi „znalezione", nie „usunięte" — niczego nie kasujemy,
        tylko zgłaszamy, że w źródle są dwa zapisy tego samego zdarzenia
        i trzeba rozstrzygnąć, który jest prawdziwy.
        """
        wlaczone = bool(self.var_zapisz_duplikaty.get())
        dane = ustawienia.wczytaj(self.katalog_danych)
        if dane.get(ustawienia.KLUCZ_ZAPISU_DUPLIKATOW) != wlaczone:
            dane[ustawienia.KLUCZ_ZAPISU_DUPLIKATOW] = wlaczone
            ustawienia.zapisz(self.katalog_danych, dane)

        duplikaty = wynik.get("duplikaty", [])
        if not wlaczone or not duplikaty:
            return []

        numery = {id(d): d["wiersz"].numer_wiersza for d in duplikaty}
        surowe = {w.get(KLUCZ_NUMERU_WIERSZA): w for w in self.surowe}
        pozycje = [{
            "numer_wiersza": numery[id(d)],
            "dane": {k: v for k, v in (surowe.get(numery[id(d)]) or {}).items()
                     if k != KLUCZ_NUMERU_WIERSZA},
            "istniejace": _po_polsku(d.get("istniejace") or {}),
            "roznice": [_ETYKIETY_POL.get(p, p) for p in d.get("roznice") or []],
        } for d in duplikaty]

        sciezka = katalog / f"{rdzen}-znalezione-duplikaty.xlsx"
        eksport.zapisz_duplikaty(sciezka, naglowki, pozycje)
        return [sciezka]

    def _zapisz_do_poprawy(self, wynik):
        """Zapisuje plik-resztę i wykaz obok pliku ŹRÓDŁOWEGO.

        Obok źródła, nie w katalogu danych aplikacji: użytkownik wie,
        gdzie położył swój Excel, a `%LOCALAPPDATA%` jest dla niego
        miejscem, którego nie znajdzie. Zwraca listę ścieżek albo pustą
        listę, gdy wszystko weszło.
        """
        if self.sciezka_pliku is None:
            return []

        katalog = self.sciezka_pliku.parent
        rdzen = self.sciezka_pliku.stem
        naglowki = list(self.surowe[0]) if self.surowe else []
        naglowki = [n for n in naglowki if n != KLUCZ_NUMERU_WIERSZA]

        pliki_duplikatow = self._zapisz_duplikaty(
            wynik, katalog, rdzen, naglowki)

        pozycje = zbierz_odrzucone(self.odrzucone, wynik.get("wymagajace_uwagi", []))
        if not pozycje:
            return pliki_duplikatow

        sciezka_reszty = katalog / f"{rdzen}-do-poprawy.xlsx"
        eksport.zapisz_niezaimportowane(
            sciezka_reszty, naglowki,
            wybierz_niezaimportowane(self.surowe, pozycje),
            powody={p["numer_wiersza"]: p["powod"] for p in pozycje},
        )
        sciezka_wykazu = katalog / f"{rdzen}-odrzucone.xlsx"
        eksport.zapisz_odrzucone(sciezka_wykazu, pozycje)
        return [sciezka_reszty, sciezka_wykazu]


class ZakladkaImportExport(ttk.Frame):
    def __init__(self, parent, conn, katalog_danych, on_zaimportowano=None,
                 autor_id=None, sesja_uuid=None, conn_rejonarz=None):
        super().__init__(parent)
        self.conn = conn
        self.conn_rejonarz = conn_rejonarz
        self.katalog_danych = katalog_danych
        self.on_zaimportowano = on_zaimportowano
        self.autor_id = autor_id
        self.sesja_uuid = sesja_uuid

        ramka_import = ttk.LabelFrame(self, text="Import z .xlsx", padding=10)
        ramka_import.pack(fill="x", padx=10, pady=10)
        ttk.Button(ramka_import, text="Wybierz plik i importuj", command=self.importuj).pack(anchor="w")
        self.etykieta_import = ttk.Label(ramka_import, text="")
        self.etykieta_import.pack(anchor="w", pady=(6, 0))

        ramka_rejonarz = ttk.LabelFrame(
            self, text="Dane referencyjne z BaŚKi", padding=10)
        ramka_rejonarz.pack(fill="x", padx=10, pady=(0, 10))
        ttk.Button(
            ramka_rejonarz, text="Wczytaj eksport z BaŚKi...",
            command=self.importuj_rejonarz,
        ).pack(anchor="w")
        self.etykieta_rejonarz = ttk.Label(ramka_rejonarz, text="")
        self.etykieta_rejonarz.pack(anchor="w", pady=(6, 0))
        self._odswiez_stan_rejonarza()

        ramka_export = ttk.LabelFrame(self, text="Export do .xlsx", padding=10)
        ramka_export.pack(fill="x", padx=10, pady=10)

        dzis = date.today()
        ttk.Label(ramka_export, text="Rok:").grid(row=0, column=0, sticky="w")
        self.var_rok = tk.IntVar(value=dzis.year)
        ttk.Entry(ramka_export, textvariable=self.var_rok, width=6).grid(row=0, column=1, padx=6)

        ttk.Label(ramka_export, text="Miesiąc:").grid(row=0, column=2, sticky="w", padx=(12, 0))
        self.var_miesiac = tk.StringVar(value=eksport.NAZWY_MIESIECY_PL[dzis.month - 1])
        ttk.Combobox(
            ramka_export, textvariable=self.var_miesiac,
            values=eksport.NAZWY_MIESIECY_PL, width=12, state="readonly",
        ).grid(row=0, column=3, padx=6)

        ttk.Button(ramka_export, text="Eksportuj...", command=self.eksportuj).grid(
            row=0, column=4, padx=(12, 0)
        )
        self.etykieta_export = ttk.Label(ramka_export, text="")
        self.etykieta_export.grid(row=1, column=0, columnspan=5, sticky="w", pady=(6, 0))

    def importuj(self):
        sciezka = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not sciezka:
            return
        surowe = _wczytaj_surowe_wiersze(sciezka)
        zwalidowane, odrzucone = zwaliduj_wiersze(surowe)
        if not zwalidowane and not odrzucone:
            self.etykieta_import.configure(text="Brak wierszy do zaimportowania w tym pliku.")
            return
        propozycje = znajdz_propozycje_scalenia_kurierow(zwalidowane)
        ostrzezenia = znajdz_ostrzezenia_podobienstwa_kurierow(zwalidowane)
        DialogKorektyImportu(
            self, self.conn, self.katalog_danych, Path(sciezka).name,
            zwalidowane, odrzucone, propozycje, ostrzezenia,
            on_gotowe=self._po_imporcie,
            autor_id=self.autor_id, sesja_uuid=self.sesja_uuid,
            status_zaufania=eksport.zweryfikuj_plik(sciezka),
            sciezka_pliku=sciezka, surowe=surowe,
            szukaj_rejonarza=self._szukaj_rejonarza(),
        )

    def _po_imporcie(self, wynik):
        tekst = f"Zaimportowano {wynik['zaimportowano']} wierszy."
        if wynik["wymagajace_uwagi"]:
            tekst += f" Do przejrzenia: {len(wynik['wymagajace_uwagi'])} (konflikty PNI/adres, duplikaty)."
        pliki = wynik.get("pliki") or []
        if pliki:
            tekst += f" Plik do poprawy: {pliki[0].name} (obok pliku źródłowego)."
        self.etykieta_import.configure(text=tekst)
        if self.on_zaimportowano:
            self.on_zaimportowano()

    def _szukaj_rejonarza(self):
        """
        Migawka jako wstrzykiwana zależność kaskady dedukcji miejscowości.

        `None` zarówno gdy migawki nie ma, jak i gdy jest PUSTA - dla
        dedukcji to ta sama sytuacja, a rozróżnianie ich dawałoby ciche
        różnice zachowania między stacją bez importu a stacją, na której
        import wczytał pusty plik.
        """
        if self.conn_rejonarz is None:
            return None
        if not rejonarz.czy_dostepny(self.conn_rejonarz):
            return None
        return rejonarz.zbuduj_szukaj(self.conn_rejonarz)

    def _odswiez_stan_rejonarza(self):
        if self.conn_rejonarz is None:
            self.etykieta_rejonarz.configure(text="Niedostępny w tej sesji.")
            return
        ile = rejonarz.policz(self.conn_rejonarz)
        self.etykieta_rejonarz.configure(
            text=f"Wczytanych adresów: {ile}." if ile
            else "Brak wczytanego rejonarza - rejon będzie zostawał jako „???”.")

    def importuj_rejonarz(self):
        """
        JEDEN przycisk na oba eksporty z BaŚKi - rodzaj rozpoznaje
        `rejonarz.wczytaj` po zawartości pliku.

        Rozróżnienie w interfejsie jest czysto funkcjonalne: użytkownik
        wskazuje plik, program ustala, czym on jest. Osobny przycisk na
        każdy rodzaj byłby przerzuceniem na niego decyzji, której nie ma
        jak podjąć - nazwy plików z BaŚKi nic o rodzaju nie mówią.

        Każdy import PODMIENIA swój rejestr w całości (migawka stanu, nie
        dziennik przyrostowy), ale rejestry są rozdzielne, więc wczytanie
        punktów nie kasuje rejonarza adresowego.
        """
        if self.conn_rejonarz is None:
            return
        sciezka = filedialog.askopenfilename(
            # OBA formaty: eksporty z BaŚKi przychodzą jako `.xls` (stary
            # OLE2), nasze własne pliki jako `.xlsx`. Rozszerzenie i tak
            # nie rozstrzyga - `arkusze.otworz` patrzy na zawartość.
            filetypes=[("Excel", "*.xls *.xlsx"), ("Wszystkie pliki", "*.*")])
        if not sciezka:
            return
        try:
            wczytane = rejonarz.wczytaj(self.conn_rejonarz, sciezka)
        except (rejonarz.NiezgodnyArkusz, arkusze.NieznanyFormat) as e:
            # Komunikat mówi WPROST, czego zabrakło, zamiast zostawiać
            # użytkownika z "nie zadziałało".
            messagebox.showerror("Nie rozpoznano pliku", str(e), parent=self)
            self.etykieta_rejonarz.configure(text="Nie wczytano - patrz komunikat.")
            return

        self.etykieta_rejonarz.configure(text=_podsumowanie_wczytania(wczytane))

    def eksportuj(self):
        try:
            rok = int(self.var_rok.get())
            miesiac = eksport.NAZWY_MIESIECY_PL.index(self.var_miesiac.get()) + 1
        except ValueError:
            self.etykieta_export.configure(text="Nieprawidłowy rok.")
            return

        sciezka = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=f"zpo-{rok}-{miesiac:02d}.xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not sciezka:
            return
        liczba = eksport.eksportuj_miesiac(self.conn, rok, miesiac, sciezka)
        self.etykieta_export.configure(text=f"Wyeksportowano {liczba} transakcji do {sciezka}.")
