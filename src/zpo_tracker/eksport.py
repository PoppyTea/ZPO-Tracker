"""
Export transakcji do pliku .xlsx o układzie identycznym ze snapshotem
źródłowym: te same nagłówki i kolejność kolumn, jeden arkusz na miesiąc,
nazwany po polsku (docs/domain-model.md: "każdy nowy miesiąc to nowy
arkusz"). Typy komórek kanonicznie czyste (int/date) - świadomie NIE
odtwarzają niespójności ręcznego wpisywania w źródle (część ilości tam jest
zapisana jako tekst), bo czyszczenie tego jest celem narzędzia.

**PNI jest wyjątkiem od tej reguły i zostaje TEKSTEM** (0.1-alpha.3.2): to
klucz tożsamości punktu (`punkty.pni_zpo UNIQUE`), a nie wielkość liczbowa.
Rzutowanie na int zamieniało "007" w 7, reimport czytał "7" i ten sam
fizyczny punkt dostawał dwa różne klucze - samo-zadana korupcja, bez udziału
jakiegokolwiek obcego pliku.

Znacznik pochodzenia + odcisk palca (0.1-alpha.3.2): plik dostaje dwie
właściwości niestandardowe dokumentu - stały znacznik "to nasz eksport"
i SHA-256 kanonicznej postaci danych. Import po nich rozpoznaje, czy
plikowi wolno ufać (patrz `zweryfikuj_plik` i import_orchestrator.py).
Sam znacznik nie wystarcza: pliki .xlsx są trywialnie edytowalne w Excelu,
więc dopiero zgodny odcisk dowodzi, że zawartość jest ta, którą zapisaliśmy.
"""
import calendar
import hashlib
from datetime import date, datetime

import openpyxl
from openpyxl.packaging.custom import StringProperty

NAZWA_ZNACZNIKA = "zpo_tracker_eksport"
NAZWA_ODCISKU = "zpo_tracker_odcisk"
WERSJA_ZNACZNIKA = "1"

# wyniki `zweryfikuj_plik`
PLIK_ZAUFANY = "zaufany"            # nasz eksport, odcisk się zgadza
PLIK_OBCY = "obcy"                  # brak znacznika - obcy/ręcznie robiony plik
PLIK_ZMODYFIKOWANY = "zmodyfikowany"  # nasz znacznik, ale zawartość już nie ta

# Dokładne stringi ze snapshotu źródłowego (białe znaki są częścią danych).
NAGLOWKI = [
    "data",
    " Pełna Nazwa Nadawcy",
    "Adres odbioru dla wszystkich nadawców",
    "Kurier",
    "Rejon",
    " Wpisujemy łączną liczbę odebranych Pocztexów",
    " Wpisujemy   w tym liczbę z Zewnetrznych Punktów Odbiorów ",
    "PNI ZPO",
    "Wpisujemy w tym liczbę odebranych z ZPO  w ramach             e Commerce -Vinted",
    "w tym Liczba z Automatów ",
    "w tym Kurier 48",
    "Paczki nierozliczone - niezrealizowane odbiory",
    "Wykonawca",
]

NAZWY_MIESIECY_PL = [
    "Styczeń", "Luty", "Marzec", "Kwiecień", "Maj", "Czerwiec",
    "Lipiec", "Sierpień", "Wrzesień", "Październik", "Listopad", "Grudzień",
]


def nazwa_arkusza(rok, miesiac):
    return NAZWY_MIESIECY_PL[miesiac - 1]


def _jako_tekst_lub_none(v):
    """PNI zawsze jako tekst - patrz docstring modułu."""
    if v is None or v == "":
        return None
    return str(v).strip() or None


def _kanoniczna_komorka(v):
    """
    Jedna komórka -> stabilny tekst do policzenia odcisku. Musi dać ten sam
    wynik po stronie ZAPISU (wartości prosto z bazy) i ODCZYTU (to, co
    openpyxl przeczyta z pliku) - stąd sprowadzenie daty/czasu do ISO daty
    (openpyxl czyta daty jako `datetime`, my zapisujemy `date`) i liczb
    całkowitych do jednej postaci (int zapisany w xlsx wraca jako int, ale
    część z nich potrafi wrócić jako float, np. 3 -> 3.0).
    """
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def odcisk_wierszy(wiersze):
    """
    SHA-256 kanonicznej postaci danych arkusza. `wiersze`: iterowalne krotek
    wartości komórek (dokładnie to, co zwraca `ws.iter_rows(values_only=True)`),
    razem z wierszem nagłówków - zmiana nagłówka też musi unieważnić odcisk.
    Separatory (`\\x1f` między komórkami, `\\x1e` między wierszami) nie mogą
    wystąpić w danych, więc "a|b" i "a", "b" dają różne odciski.
    """
    tresc = "\x1e".join(
        "\x1f".join(_kanoniczna_komorka(k) for k in wiersz) for wiersz in wiersze
    )
    return hashlib.sha256(tresc.encode("utf-8")).hexdigest()


def zweryfikuj_plik(sciezka):
    """
    PLIK_ZAUFANY / PLIK_OBCY / PLIK_ZMODYFIKOWANY - patrz stałe wyżej i
    docstring modułu. Nieczytelny plik to PLIK_OBCY, nie wyjątek: decyzja
    o zaufaniu nigdy nie może wysadzić importu, a "nie umiem tego
    zweryfikować" znaczy dokładnie tyle co "nie ufam".
    """
    try:
        wb = openpyxl.load_workbook(sciezka, data_only=True)
    except Exception:
        return PLIK_OBCY

    wlasciwosci = {p.name: p.value for p in wb.custom_doc_props.props}
    if NAZWA_ZNACZNIKA not in wlasciwosci:
        return PLIK_OBCY

    ws = wb[wb.sheetnames[0]]
    if wlasciwosci.get(NAZWA_ODCISKU) != odcisk_wierszy(ws.iter_rows(values_only=True)):
        return PLIK_ZMODYFIKOWANY
    return PLIK_ZAUFANY


def pobierz_transakcje_miesiaca(conn, rok, miesiac):
    """Transakcje z danego miesiąca (włącznie z rzadkimi ilościami)."""
    pierwszy = date(rok, miesiac, 1)
    ostatni = date(rok, miesiac, calendar.monthrange(rok, miesiac)[1])
    wiersze = conn.execute(
        """SELECT t.data, n.nazwa AS nadawca, a.surowy AS adres,
                  k.imie_nazwisko AS kurier,
                  r.kod AS rejon, t.ilosc_total, t.ilosc_zpo, p.pni_zpo,
                  t.ilosc_vinted, t.ilosc_automaty, t.ilosc_kurier48,
                  t.ilosc_niezrealizowane, w.nazwa AS wykonawca
           FROM transakcje t
           JOIN kurierzy k ON k.id = t.kurier_id
           JOIN punkty p ON p.id = t.punkt_id
           JOIN nadawcy n ON n.id = p.nadawca_id
           JOIN adresy a ON a.id = p.adres_id
           LEFT JOIN rejony r ON r.id = t.rejon_id
           LEFT JOIN wykonawcy w ON w.id = t.wykonawca_id
           WHERE t.data BETWEEN ? AND ?
           ORDER BY t.data, t.id""",
        (pierwszy.isoformat(), ostatni.isoformat()),
    ).fetchall()
    return [dict(w) for w in wiersze]


def eksportuj_miesiac(conn, rok, miesiac, sciezka):
    """
    Zapisuje transakcje z danego miesiąca do pliku .xlsx - jeden arkusz
    nazwany miesiącem, kolumny A-M identyczne ze snapshotem źródłowym.
    Zwraca liczbę wyeksportowanych wierszy.
    """
    wiersze = pobierz_transakcje_miesiaca(conn, rok, miesiac)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = nazwa_arkusza(rok, miesiac)
    ws.append(NAGLOWKI)

    for w in wiersze:
        data_val = w["data"]
        if isinstance(data_val, str):
            data_val = date.fromisoformat(data_val)
        ws.append([
            data_val,
            w["nadawca"],
            w["adres"],
            w["kurier"],
            w["rejon"],
            w["ilosc_total"],
            w["ilosc_zpo"],
            _jako_tekst_lub_none(w["pni_zpo"]),
            w["ilosc_vinted"],
            w["ilosc_automaty"],
            w["ilosc_kurier48"],
            w["ilosc_niezrealizowane"],
            w["wykonawca"],
        ])

    for komorka in ws["A"][1:]:
        komorka.number_format = "yyyy-mm-dd"

    # znacznik + odcisk liczone z GOTOWEGO arkusza, nie z `wiersze` z bazy:
    # import policzy hash dokładnie z tych samych komórek, więc obie strony
    # muszą wyjść od tej samej reprezentacji (patrz docstring modułu)
    wb.custom_doc_props.append(
        StringProperty(name=NAZWA_ZNACZNIKA, value=WERSJA_ZNACZNIKA))
    wb.custom_doc_props.append(
        StringProperty(name=NAZWA_ODCISKU,
                        value=odcisk_wierszy(ws.iter_rows(values_only=True))))

    wb.save(sciezka)
    return len(wiersze)


# --- raport odrzuconych wierszy importu (0.1-alpha.3.3) -----------------

NAGLOWEK_NUMERU = "Wiersz w pliku"
NAGLOWEK_POWODU = "Powód odrzucenia"


def zapisz_odrzucone(sciezka, pozycje):
    """
    Zapisuje `.xlsx` z wykazem wierszy, które nie weszły do bazy.

    Sens tego pliku: opcja "pomiń niespójności" bez wykazu tego, co
    pominięto, jest cichą utratą danych z ładniejszą nazwą. Z wykazem -
    i z NUMEREM WIERSZA - staje się listą zadań do poprawienia
    w źródłowym Excelu.

    Kolumny: numer wiersza, powód, a dalej oryginalne nagłówki źródła.
    Kolejność kolumn danych idzie za `NAGLOWKI` (czyli tak jak w pliku
    źródłowym), a nagłówki spoza tej listy dopisują się na końcu -
    odrzucenia z dwóch etapów mają różne kształty i raport musi pomieścić
    oba, nie zgubić kolumn drugiego.

    Plik powstaje także dla pustej listy: "nic nie odrzucono" ma być
    widocznym artefaktem, nie brakiem pliku, którego nie wiadomo jak
    zinterpretować. Zwraca liczbę wypisanych wierszy.
    """
    kolumny_danych = []
    for pozycja in pozycje:
        for klucz in pozycja["dane"]:
            if klucz not in kolumny_danych:
                kolumny_danych.append(klucz)
    kolumny_danych.sort(key=lambda k: (NAGLOWKI.index(k) if k in NAGLOWKI else len(NAGLOWKI)))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Odrzucone"
    ws.append([NAGLOWEK_NUMERU, NAGLOWEK_POWODU] + kolumny_danych)

    for pozycja in pozycje:
        komorki = [pozycja.get("numer_wiersza"), pozycja.get("powod")]
        for klucz in kolumny_danych:
            komorki.append(pozycja["dane"].get(klucz))
        ws.append(komorki)
        # PNI zostaje TEKSTEM - ta sama pułapka co w eksporcie miesiąca:
        # "007" rzutowane na liczbę to "7", czyli inny punkt.
        if "PNI ZPO" in kolumny_danych:
            kolumna = kolumny_danych.index("PNI ZPO") + 3
            ws.cell(row=ws.max_row, column=kolumna).number_format = "@"

    wb.save(sciezka)
    return len(pozycje)


def zapisz_niezaimportowane(sciezka, naglowki, wiersze, powody=None, z_powodem=True):
    """
    Zapisuje kopię pliku źródłowego POZBAWIONĄ wierszy, które weszły.

    Inny artefakt niż `zapisz_odrzucone` i do innego celu: tamten jest
    wykazem do czytania, ten jest plikiem do PRACY. Poprawia się go
    i importuje ponownie, zamiast wyłuskiwać wiersze z raportu.

    Dlatego struktura jest wierna oryginałowi: te same nagłówki, ta sama
    kolejność, te same wartości. `powody` (numer wiersza -> tekst) dopisują
    się jako OSTATNIA kolumna - przy ponownym imporcie i tak zostanie
    zignorowana, bo mapowanie filtruje po znanych nagłówkach, a bez niej
    użytkownik nie wie, co właściwie ma poprawić. `z_powodem=False` daje
    kopię co do kolumny.
    """
    from zpo_tracker.import_orchestrator import KLUCZ_NUMERU_WIERSZA

    powody = powody or {}
    kolumny = list(naglowki)
    if z_powodem:
        kolumny.append(NAGLOWEK_POWODU)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Do poprawy"
    ws.append(kolumny)

    for wiersz in wiersze:
        komorki = [wiersz.get(naglowek) for naglowek in naglowki]
        if z_powodem:
            komorki.append(powody.get(wiersz.get(KLUCZ_NUMERU_WIERSZA)))
        ws.append(komorki)
        if "PNI ZPO" in naglowki:
            kolumna = naglowki.index("PNI ZPO") + 1
            ws.cell(row=ws.max_row, column=kolumna).number_format = "@"

    wb.save(sciezka)
    return len(wiersze)


# --- znalezione duplikaty ------------------------------------------------

OPIS_BEZ_ROZNIC = "identyczne - czysta kopia"
NAGLOWEK_ROZNIC = "Czym się różnią"
NAGLOWEK_WERSJI = "Wersja"
WERSJA_W_BAZIE = "już w bazie"
WERSJA_Z_PLIKU = "z importowanego pliku"


def zapisz_duplikaty(sciezka, naglowki, pozycje):
    """
    Zapisuje `.xlsx` z wierszami, które odpadły jako duplikat klucza
    (data + kurier + punkt), pokazując OBIE WERSJE obok siebie.

    Osobny artefakt od `zapisz_niezaimportowane`, i to jest sedno
    zgłoszenia: tamten plik jest DO PRACY - poprawia się go i wczytuje
    ponownie. Duplikat w nim albo wraca jako ten sam duplikat, albo
    trzeba go ręcznie skasować; jedno i drugie to strata czasu, bo
    w duplikacie nie ma czego poprawić.

    Ten plik jest do ROZSTRZYGNIĘCIA, nie do poprawiania. Pomiar na
    realnym sierpniu pokazał, dlaczego to nie to samo: z 30 powtórzonych
    kluczy 19 RÓŻNI SIĘ ilością, „w tym ZPO", PNI, wykonawcą albo
    rejonem, a tylko 8 sąsiaduje ze sobą (mediana odległości 4 wiersze,
    maksimum 3431). W arkuszu sklejonym z pracy sześciu osób znaczy to,
    że ten sam odbiór wpisały DWIE OSOBY w dwóch miejscach - i to
    człowiek musi powiedzieć, która wersja jest prawdziwa.

    Stąd dwa wiersze na każdy przypadek (stan z bazy i wiersz z pliku)
    oraz kolumna mówiąca, CZYM się różnią: przy trzydziestu przypadkach
    szukanie różnicy wzrokiem po kilkunastu kolumnach to gwarancja
    przeoczenia.

    Pusta lista NIE tworzy pliku - inaczej niż wykaz odrzuconych, gdzie
    pusty plik znaczy „sprawdziłem, nic nie odrzucono". Tutaj istnienie
    pliku samo w sobie znaczy „masz coś do zrobienia", więc pusty byłby
    fałszywym alarmem. Zwraca liczbę zgłoszonych przypadków.
    """
    if not pozycje:
        return 0

    kolumny = [n for n in naglowki if n]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Znalezione duplikaty"
    ws.append([NAGLOWEK_NUMERU, NAGLOWEK_WERSJI, NAGLOWEK_ROZNIC] + kolumny)

    for pozycja in pozycje:
        dane = pozycja.get("dane") or {}
        istniejace = pozycja.get("istniejace") or {}
        roznice = pozycja.get("roznice") or []
        opis = ", ".join(roznice) if roznice else OPIS_BEZ_ROZNIC

        # Najpierw stan z bazy, potem wiersz z pliku - w tej kolejności,
        # bo pytanie brzmi "czy podmienić to, co mam, na to, co przyszło".
        ws.append([pozycja.get("numer_wiersza"), WERSJA_W_BAZIE, opis]
                  + [istniejace.get(k, dane.get(k)) for k in kolumny])
        ws.append([pozycja.get("numer_wiersza"), WERSJA_Z_PLIKU, opis]
                  + [dane.get(k) for k in kolumny])

        if "PNI ZPO" in kolumny:
            # PNI zostaje TEKSTEM - ta sama pułapka co w eksporcie
            # miesiąca: "007" rzutowane na liczbę to "7", czyli inny punkt.
            kolumna = kolumny.index("PNI ZPO") + 4
            for offset in (0, 1):
                ws.cell(row=ws.max_row - offset, column=kolumna).number_format = "@"

    wb.save(sciezka)
    return len(pozycje)
