"""
Jedna warstwa czytania skoroszytów, niezależna od formatu.

Powód istnienia jest twardy: `openpyxl` **nie otwiera `.xls` w ogóle** -
to inny format pliku (OLE2), nie kwestia wersji biblioteki. Wszystkie
eksporty z BaŚKi przychodzą właśnie w nim, więc bez tej warstwy nie da
się wczytać ani rejonarza, ani rejestru punktów ZPO.

Dwie decyzje, na których stoi ten moduł:

**Format rozpoznajemy po zawartości, nie po rozszerzeniu.** Pliki
przychodzą przez przeglądarkę i bywają przemianowane - rozszerzenie jest
deklaracją użytkownika, magiczne bajty są faktem.

**Oba silniki oddają te same wartości.** `xlrd` zwraca każdą liczbę jako
`float`, więc numer domu `8` przychodzi jako `8.0` i po sklejeniu klucza
daje `"8.0"` - czyli inny adres niż `"8"`. Ujednolicenie siedzi tutaj,
żeby żaden konsument nie musiał o tym pamiętać; pamiętanie o takiej
rzeczy w pięciu miejscach kończy się zapomnieniem w szóstym.
"""
from pathlib import Path

# Magiczne bajty. OLE2 to stary binarny Excel (.xls), ZIP to nowy (.xlsx,
# który jest archiwum z XML-ami w środku).
_OLE2 = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
_ZIP = b"PK\x03\x04"


class NieznanyFormat(Exception):
    """Plik nie jest ani `.xls`, ani `.xlsx` - niezależnie od nazwy."""


def otworz(sciezka):
    """
    Otwiera skoroszyt, rozpoznając format po pierwszych bajtach.

    Zwraca obiekt z `nazwy_arkuszy()`, `wiersze(nazwa)` i `zamknij()`,
    działający też jako menedżer kontekstu.
    """
    sciezka = Path(sciezka)
    with open(sciezka, "rb") as f:
        naglowek = f.read(8)

    if naglowek.startswith(_OLE2):
        return _SkoroszytXls(sciezka)
    if naglowek.startswith(_ZIP):
        return _SkoroszytXlsx(sciezka)
    raise NieznanyFormat(
        f"Plik {sciezka.name} nie jest arkuszem Excela (ani .xls, ani .xlsx). "
        f"Sprawdź, czy to na pewno eksport, a nie np. strona logowania "
        f"zapisana przez przeglądarkę."
    )


def _ujednolic(wartosc):
    """
    `8.0` -> `8`, reszta bez zmian.

    `xlrd` nie odróżnia liczby całkowitej od ułamkowej - w formacie `.xls`
    każda liczba jest zmiennoprzecinkowa. `openpyxl` oddaje `int`, więc
    bez tego ten sam numer domu miałby dwie różne postaci zależnie od
    tego, w jakim formacie przyszedł plik.
    """
    if isinstance(wartosc, float) and wartosc.is_integer():
        return int(wartosc)
    return wartosc


class _Skoroszyt:
    def __enter__(self):
        return self

    def __exit__(self, *_wyjatek):
        self.zamknij()
        return False

    def zwolnione(self):
        """Arkusze zwolnione z pamięci. Puste dla silników, które i tak
        strumieniują - patrz `_SkoroszytXls.zwolnione`."""
        return []


class _SkoroszytXlsx(_Skoroszyt):
    """
    PUŁAPKA: `openpyxl.load_workbook` sprawdza ROZSZERZENIE ścieżki i
    odmawia otwarcia pliku nazwanego `.xls`, nawet gdy jego zawartość jest
    poprawnym `.xlsx`. Robi więc dokładnie to, czego ten moduł ma nie
    robić - ufa nazwie zamiast bajtom.

    Podanie otwartego uchwytu zamiast ścieżki omija tę kontrolę: obiekt
    plikowy nie ma rozszerzenia, więc `openpyxl` czyta zawartość. Uchwyt
    musi żyć aż do `zamknij()`, bo tryb `read_only` strumieniuje z pliku
    i zamknięcie go wcześniej urwałoby odczyt w połowie.
    """

    def __init__(self, sciezka):
        import openpyxl
        self._uchwyt = open(sciezka, "rb")
        self._wb = openpyxl.load_workbook(
            self._uchwyt, read_only=True, data_only=True)

    def nazwy_arkuszy(self):
        return list(self._wb.sheetnames)

    def wiersze(self, nazwa):
        for wiersz in self._wb[nazwa].iter_rows(values_only=True):
            yield tuple(_ujednolic(w) for w in wiersz)

    def zamknij(self):
        self._wb.close()
        self._uchwyt.close()


class _SkoroszytXls(_Skoroszyt):
    """
    `xlrd` wczytuje CAŁY arkusz do pamięci - nie ma trybu strumieniowego.

    Przy realnym rejonarzu (219 arkuszy, 21 MB, 277 tys. wierszy) trzymanie
    wszystkich naraz jest jedynym powodem, dla którego ten import mógłby
    się nie zmieścić na słabszej stacji. Stąd `on_demand=True` i jawne
    `unload_sheet` po każdym przejściu.

    Zwolnienie jest niewidoczne dla wołającego: `xlrd` wczyta arkusz
    ponownie, jeśli ktoś po niego wróci.
    """

    def __init__(self, sciezka):
        import xlrd
        self._wb = xlrd.open_workbook(str(sciezka), on_demand=True)
        self._zwolnione = []

    def nazwy_arkuszy(self):
        return list(self._wb.sheet_names())

    def wiersze(self, nazwa):
        arkusz = self._wb.sheet_by_name(nazwa)
        try:
            for i in range(arkusz.nrows):
                yield tuple(_ujednolic(w) for w in arkusz.row_values(i))
        finally:
            # W `finally`, bo konsument bywa leniwy i może przerwać
            # iterację w połowie (np. `break` po znalezieniu nagłówka) -
            # arkusz ma zostać zwolniony także wtedy.
            self._wb.unload_sheet(nazwa)
            if nazwa not in self._zwolnione:
                self._zwolnione.append(nazwa)

    def zwolnione(self):
        return list(self._zwolnione)

    def zamknij(self):
        self._wb.release_resources()
